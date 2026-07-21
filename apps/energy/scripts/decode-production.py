"""O2 · Production decoder.

Parses data-energy/raw/Production_data/Volve production data.xlsx (both the
Daily and Monthly sheets), preserving exact column names. Emits:
  - data-energy/interim/production-daily.json     (raw rows, columns untouched)
  - data-energy/interim/production-monthly.json
  - data-energy/processed/production.json          (canonical per-row + summary)
  - data-energy/processed/production.csv           (canonical daily, flat)
  - docs/arganta-energy/qc/production.md

dataNature = 'reported' (allocated/reported production volumes, not measured).
No unit conversion. Source units (Sm3, bar, etc.) preserved as given.
"""
from __future__ import annotations
import csv, os, sys, datetime
import openpyxl
sys.path.insert(0, os.path.dirname(__file__))
from _evidence import DATA_ROOT, RAW, evidence, write_json, log

XLSX = os.path.join(RAW, "Production_data", "Volve production data.xlsx")

def isodate(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v

def main():
    sid, ev = evidence(XLSX)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ---- Daily sheet ----
    ws = wb["Daily Production Data"]
    rows = ws.iter_rows(values_only=True)
    daily_header = list(next(rows))
    daily_raw = []
    for r in rows:
        if all(c is None for c in r):
            continue
        daily_raw.append({daily_header[i]: (isodate(v)) for i, v in enumerate(r)})

    # ---- Monthly sheet ----
    ws2 = wb["Monthly Production Data"]
    rows2 = list(ws2.iter_rows(values_only=True))
    monthly_header = list(rows2[0])
    monthly_units = list(rows2[1]) if len(rows2) > 1 else []
    monthly_raw = []
    for r in rows2[2:]:
        if all(c is None for c in r):
            continue
        monthly_raw.append({monthly_header[i]: v for i, v in enumerate(r)})

    write_json(os.path.join(DATA_ROOT, "interim", "production-daily.json"),
               {"source_id": sid, "sheet": "Daily Production Data",
                "columns": daily_header, "rows": daily_raw})
    write_json(os.path.join(DATA_ROOT, "interim", "production-monthly.json"),
               {"source_id": sid, "sheet": "Monthly Production Data",
                "columns": monthly_header, "units_row": monthly_units,
                "rows": monthly_raw})

    # ---- Canonical daily ProductionRecord ----
    canon = []
    for r in daily_raw:
        canon.append({
            "source_well_bore_name": r.get("NPD_WELL_BORE_NAME"),
            "well_bore_code": r.get("WELL_BORE_CODE"),
            "npd_well_bore_code": r.get("NPD_WELL_BORE_CODE"),
            "date": r.get("DATEPRD"),
            "on_stream_hrs": r.get("ON_STREAM_HRS"),
            "avg_downhole_pressure": r.get("AVG_DOWNHOLE_PRESSURE"),
            "avg_downhole_temperature": r.get("AVG_DOWNHOLE_TEMPERATURE"),
            "avg_dp_tubing": r.get("AVG_DP_TUBING"),
            "avg_annulus_press": r.get("AVG_ANNULUS_PRESS"),
            "avg_choke_size_p": r.get("AVG_CHOKE_SIZE_P"),
            "avg_choke_uom": r.get("AVG_CHOKE_UOM"),
            "avg_whp_p": r.get("AVG_WHP_P"),
            "avg_wht_p": r.get("AVG_WHT_P"),
            "dp_choke_size": r.get("DP_CHOKE_SIZE"),
            "bore_oil_vol": r.get("BORE_OIL_VOL"),
            "bore_gas_vol": r.get("BORE_GAS_VOL"),
            "bore_wat_vol": r.get("BORE_WAT_VOL"),
            "bore_wi_vol": r.get("BORE_WI_VOL"),
            "flow_kind": r.get("FLOW_KIND"),
            "well_type": r.get("WELL_TYPE"),
            "field": r.get("NPD_FIELD_NAME"),
            "facility": r.get("NPD_FACILITY_NAME"),
            "source_id": sid,
            "dataNature": "reported",
        })

    # ---- Per-wellbore summary ----
    summary = {}
    for c in canon:
        wb_name = c["source_well_bore_name"]
        s = summary.setdefault(wb_name, {
            "wellbore": wb_name, "rows": 0, "date_min": None, "date_max": None,
            "flow_kinds": set(), "well_types": set(),
            "sum_oil_sm3": 0.0, "sum_gas_sm3": 0.0, "sum_wat_sm3": 0.0, "sum_wi_sm3": 0.0,
        })
        s["rows"] += 1
        d = c["date"]
        if d:
            if s["date_min"] is None or d < s["date_min"]: s["date_min"] = d
            if s["date_max"] is None or d > s["date_max"]: s["date_max"] = d
        if c["flow_kind"]: s["flow_kinds"].add(c["flow_kind"])
        if c["well_type"]: s["well_types"].add(c["well_type"])
        for k, col in (("sum_oil_sm3","bore_oil_vol"),("sum_gas_sm3","bore_gas_vol"),
                       ("sum_wat_sm3","bore_wat_vol"),("sum_wi_sm3","bore_wi_vol")):
            v = c[col]
            if isinstance(v,(int,float)): s[k] += v
    for s in summary.values():
        s["flow_kinds"] = sorted(s["flow_kinds"]) or None
        s["well_types"] = sorted(s["well_types"]) or None

    # canonical monthly
    canon_monthly = []
    for r in monthly_raw:
        canon_monthly.append({
            "source_well_bore_name": r.get("Wellbore name"),
            "npd_code": r.get("NPDCode"),
            "year": r.get("Year"), "month": r.get("Month"),
            "on_stream_hrs": r.get("On Stream"),
            "oil_sm3": r.get("Oil"), "gas_sm3": r.get("Gas"),
            "water_sm3": r.get("Water"), "gi_sm3": r.get("GI"), "wi_sm3": r.get("WI"),
            "source_id": sid, "dataNature": "reported",
        })

    write_json(os.path.join(DATA_ROOT, "processed", "production.json"), {
        "source_id": sid, "evidence": ev,
        "daily_columns_verbatim": daily_header,
        "monthly_columns_verbatim": monthly_header,
        "monthly_units_verbatim": monthly_units,
        "daily_rows": canon,
        "monthly_rows": canon_monthly,
        "wellbore_summary": list(summary.values()),
        "units_note": "Volumes in Sm3 (standard cubic metres) per source; pressures/temps in source units; NO conversion applied.",
    })

    # flat CSV of canonical daily
    csv_path = os.path.join(DATA_ROOT, "processed", "production.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(canon[0].keys()))
        w.writeheader()
        w.writerows(canon)

    # QC report
    qc = os.path.join(DATA_ROOT, "..", "docs", "arganta-energy", "qc", "production.md")
    qc = os.path.abspath(qc)
    with open(qc, "w", encoding="utf-8") as fh:
        fh.write("# QC · Production\n\n")
        fh.write(f"- source_id: `{sid}`\n- sha256: `{ev['sha256']}`\n")
        fh.write(f"- daily rows decoded: **{len(canon)}**  ·  monthly rows: **{len(canon_monthly)}**\n")
        fh.write("- dataNature: `reported`  ·  no unit conversion applied\n\n")
        fh.write("## Daily columns (verbatim)\n\n" + ", ".join(f"`{c}`" for c in daily_header) + "\n\n")
        fh.write("## Per-wellbore summary\n\n")
        fh.write("| wellbore | rows | date range | flow kinds | well types | oil Sm3 | gas Sm3 | water Sm3 | WI Sm3 |\n")
        fh.write("|---|--:|---|---|---|--:|--:|--:|--:|\n")
        for s in sorted(summary.values(), key=lambda x: x["wellbore"] or ""):
            fh.write(f"| {s['wellbore']} | {s['rows']} | {s['date_min']} → {s['date_max']} | "
                     f"{','.join(s['flow_kinds'] or [])} | {','.join(s['well_types'] or [])} | "
                     f"{s['sum_oil_sm3']:.0f} | {s['sum_gas_sm3']:.0f} | {s['sum_wat_sm3']:.0f} | {s['sum_wi_sm3']:.0f} |\n")

    log(f"[production] daily={len(canon)} monthly={len(canon_monthly)} wellbores={len(summary)}")
    print(f"OK daily={len(canon)} monthly={len(canon_monthly)} wellbores={len(summary)}")

if __name__ == "__main__":
    main()
