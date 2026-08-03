# export_kb_markdown.py — the missing link in the chain:
#     ArgantaEnergy-Master-KB.xlsx  →  Markdown vault  →  knowledge graph
#
# Reads the master workbook and emits one Obsidian-style .md note per row, with
# YAML frontmatter and [[wikilinks]] resolved from the workbook's own foreign keys —
# so a key IS a link, matching the convention already used by
# apps/energy/src/cosmo/knowledge-model.ts.
#
# ATTRIBUTION: notes generated from the "Doust Figure Sourcing" tab always carry
# their credit line. The organization's clearance for that material is conditional on
# attribution being displayed, so it is emitted with the note, never stripped.
#
# Run:  python docs/arganta-energy/knowledge-base/export_kb_markdown.py
import os
import re

import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "ArgantaEnergy-Master-KB.xlsx")
OUT = os.path.join(HERE, "vault")

# sheet -> (folder, title column, id column, [FK column -> target sheet])
PLAN = {
    "Region":              ("01_Regions", "name", "region_id", {}),
    "Country":             ("02_Countries", "name", "country_id", {}),
    "Province":            ("03_Provinces", "name", "province_id", {"region_id": "Region"}),
    "Basin":               ("04_Basins", "name", "basin_id", {"province_id": "Province"}),
    "Basin Cycle":         ("05_BasinCycles", "title", "cycle_id", {"basin_id": "Basin"}),
    "Stratigraphic Units": ("06_StratUnits", "unit_name", "unit_name", {"cycle_id": "Basin Cycle"}),
    "Petroleum System":    ("07_PetroleumSystems", "name", "tps_id", {"province_id": "Province"}),
    "Assessment Unit":     ("08_AssessmentUnits", "name", "au_id", {"tps_id": "Petroleum System"}),
    "Play":                ("09_Plays", "name", "play_id", {"cycle_id": "Basin Cycle", "tps_id": "Petroleum System"}),
    "Opportunity":         ("10_Opportunities", "name", "opportunity_id", {"play_id": "Play"}),
    "Field":               ("11_Fields", "name", "field_id", {"basin_id": "Basin", "country_id": "Country"}),
    "Reservoir":           ("12_Reservoirs", "formation_name", "reservoir_id", {"field_id": "Field"}),
    "Well":                ("13_Wells", "well_id", "well_id", {"field_id": "Field"}),
    "Wellbore":            ("14_Wellbores", "wellbore_id", "wellbore_id", {"well_id": "Well"}),
    "Concepts":            ("15_Concepts", "title", "concept_id", {"citation_id": "Citations"}),
    "Citations":           ("16_Citations", "title", "citation_id", {}),
    "Doust Figure Sourcing": ("17_BasinFigures", "caption", "fig_no", {"citation_id": "Citations"}),
}

SAFE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def safe(name: str) -> str:
    return SAFE.sub("-", str(name)).strip().strip(".")[:120] or "untitled"


def rows_of(ws):
    """Yield real data rows only.

    Several sheets carry a merged explanatory NOTE row a couple of lines below the
    table. openpyxl reports its prose in column A, so a plain "is column A empty?"
    test lets it through and mints a junk note. A note row is recognisable as the
    one where every column but the first is empty and the first holds a long
    sentence — real rows always populate at least one further column."""
    it = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(it)]
    for row in it:
        if row[0] is None or str(row[0]).strip() == "":
            continue
        others = [c for c in row[1:] if c is not None and str(c).strip() != ""]
        if not others and len(str(row[0])) > 120:
            continue  # merged explanatory note row, not data
        yield headers, dict(zip(headers, row))


def yaml_scalar(v) -> str:
    s = str(v)
    return f'"{s}"' if re.search(r'[:#\[\]{}",\n]', s) else s


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"workbook not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

    # id -> display title, so a foreign key can resolve to a real [[note name]]
    titles = {}
    for sheet, (_folder, tcol, icol, _fks) in PLAN.items():
        if sheet not in wb.sheetnames:
            continue
        for headers, r in rows_of(wb[sheet]):
            if icol in r and tcol in r and r[icol] is not None:
                titles[str(r[icol])] = safe(r.get(tcol) or r[icol])

    # This repo lives inside OneDrive, which intermittently holds a lock on freshly
    # written folders — a blunt rmtree fails with WinError 32 partway through and
    # leaves a half-deleted vault. Delete file-by-file, tolerate what is locked, and
    # keep going; every note is rewritten below anyway.
    if os.path.isdir(OUT):
        stale = 0
        for root, _dirs, files in os.walk(OUT):
            for name in files:
                try:
                    os.remove(os.path.join(root, name))
                except OSError:
                    stale += 1
        if stale:
            print(f"  note: {stale} existing file(s) locked (OneDrive) — overwritten in place")
    written = 0
    per_sheet = {}

    for sheet, (folder, tcol, icol, fks) in PLAN.items():
        if sheet not in wb.sheetnames:
            print(f"  (skip, no sheet) {sheet}")
            continue
        d = os.path.join(OUT, folder)
        os.makedirs(d, exist_ok=True)
        n = 0
        seen = set()
        for headers, r in rows_of(wb[sheet]):
            title = safe(r.get(tcol) or r.get(icol))
            stem = title
            k = 2
            while stem.lower() in seen:           # two fields can share a name
                stem = f"{title} ({k})"; k += 1
            seen.add(stem.lower())

            fm = [f"entity: {sheet}"]
            for col, val in r.items():
                if val is None or str(val).strip() == "" or col == "":
                    continue
                if col in fks:
                    target = titles.get(str(val))
                    fm.append(f"{col}: {yaml_scalar(f'[[{target}]]' if target else val)}")
                else:
                    fm.append(f"{col}: {yaml_scalar(val)}")

            body = [f"# {r.get(tcol) or r.get(icol)}", ""]
            for col in fks:
                if r.get(col) and titles.get(str(r[col])):
                    body.append(f"- **{col.replace('_id','').replace('_',' ').title()}:** [[{titles[str(r[col])]}]]")
            if body[-1] != "":
                body.append("")
            for col, val in r.items():
                if col in ("notes", "description", "body", "essential_elements_note", "caption") and val:
                    body += [f"## {col.replace('_',' ').title()}", str(val), ""]

            # Attribution is a CONDITION of use for figure notes — never omit it.
            if sheet == "Doust Figure Sourcing":
                src, sourcing = r.get("source_short"), r.get("sourcing")
                credit = (f'© H. Doust — Doust, H., "Dissecting Sedimentary Basins", fig. {r.get("fig_no")}'
                          if sourcing == "Own" else
                          f'After {src} — reproduced in Doust, H., "Dissecting Sedimentary Basins", fig. {r.get("fig_no")}')
                body += ["## Attribution", credit,
                         "", "Cleared for internal scientific/educational use with attribution; "
                         "not cleared for public redistribution.", ""]
                if r.get("image_file"):
                    body += [f"![[{os.path.basename(str(r['image_file']))}]]", ""]

            with open(os.path.join(d, f"{stem}.md"), "w", encoding="utf-8") as f:
                f.write("---\n" + "\n".join(fm) + "\n---\n\n" + "\n".join(body))
            n += 1
        per_sheet[sheet] = n
        written += n
        print(f"  {sheet:24s} -> {folder:22s} {n:5d} notes")

    with open(os.path.join(OUT, "README.md"), "w", encoding="utf-8") as f:
        f.write("# ArgantaEnergy Knowledge Vault\n\nGenerated from `ArgantaEnergy-Master-KB.xlsx` by "
                "`export_kb_markdown.py` — do not hand-edit; edit the workbook and re-run.\n\n"
                + "\n".join(f"- **{k}** — {v} notes" for k, v in per_sheet.items())
                + "\n\nFigure notes carry a required attribution line. The image files they reference "
                  "live in `apps/energy/public/doust-figures/` and are gitignored: cleared for internal "
                  "scientific/educational use with attribution, not for public redistribution.\n")
    print(f"\n{written} notes -> {OUT}")


if __name__ == "__main__":
    main()
