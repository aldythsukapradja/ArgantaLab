# build_master_kb.py — builds the ArgantaEnergy Master Knowledge Base workbook.
# Every row is either (a) real, already-shipped data from apps/energy/src (ATLAS,
# knowledge-model.ts, explData.ts, wb/index.json, fieldcraft/catalog.ts) or
# (b) independently-verified USGS research from this session, or (c) Doust's own
# genuinely-original thesis/criteria/citations. This workbook names NO confidential
# external project anywhere (not even to say it was excluded) — the Report Sections
# tab reflects only generic, publicly-describable Play-Based-Exploration methodology
# (CRS/CCRS, GCF, GRV two-surface) that is standard industry practice.
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FINAL_OUT = r"C:\Users\aldhy\OneDrive\Documents\GitHub\ArgantaLab\docs\arganta-energy\knowledge-base\ArgantaEnergy-Master-KB.xlsx"
OUT = r"C:\Users\aldhy\AppData\Local\Temp\claude\C--Users-aldhy-OneDrive-Documents-GitHub-ArgantaLab\3d65e7cb-77e7-40b1-91c7-498b0b57e202\scratchpad\ArgantaEnergy-Master-KB.xlsx"
WORLD = json.load(open(r"C:\Users\aldhy\AppData\Local\Temp\claude\C--Users-aldhy-OneDrive-Documents-GitHub-ArgantaLab\3d65e7cb-77e7-40b1-91c7-498b0b57e202\scratchpad\world-kb.json", encoding="utf-8"))
print("WORLD data:", WORLD["meta"]["counts"])

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F2937")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")
BODY_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="9CA3AF")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# dataNature -> light fill, matching the app's NatureBadge accent family
NATURE_FILL = {
    "measured": "E6F7F5", "reported": "FFF3D6", "interpreted": "FDE9D9",
    "derived": "F0E6FB", "forecast": "E8F0FE", "scenario": "FCE8EC", "reference": "EDEFF2",
}

wb = openpyxl.Workbook()
wb.remove(wb.active)

# last real DATA row (excl. header, excl. any note rows below) per sheet, captured as
# each table is built — used at the very end for bounded, note-row-safe COUNTA ranges.
LAST_ROW = {}


def add_sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def write_table(ws, headers, rows, start_row=1, nature_col=None, freeze=True, col_widths=None):
    """Write a header row + data rows with consistent styling. nature_col = 0-based
    index of a dataNature column, used to tint that cell by its value."""
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=j + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=False)
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            c = ws.cell(row=start_row + 1 + i, column=j + 1, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if nature_col is not None and j == nature_col and isinstance(v, str) and v in NATURE_FILL:
                c.fill = PatternFill("solid", fgColor=NATURE_FILL[v])
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    widths = col_widths or [18] * len(headers)
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.row_dimensions[start_row].height = 20
    return start_row + 1 + len(rows)


def write_table_fast(ws, headers, rows, start_row=1, freeze=True, col_widths=None):
    """Like write_table but WITHOUT per-cell borders/wrap — for large bulk tables
    (thousands of rows) where per-cell Border objects would be slow. Still gets the
    same header style and Arial body font."""
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=j + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            ws.cell(row=start_row + 1 + i, column=j + 1, value=v).font = BODY_FONT
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    widths = col_widths or [18] * len(headers)
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w
    ws.row_dimensions[start_row].height = 20
    return start_row + 1 + len(rows)


import re


def slug(s):
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", s.lower()))


def note_row(ws, row, text, span):
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


# ══════════════════════════════════════════════════════════════════════════════
# README
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("README")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 90
ws["A1"] = "ArgantaEnergy Master Knowledge Base"
ws["A1"].font = TITLE_FONT
ws["A2"] = "The single source-of-truth workbook: entities → Markdown → Knowledge Graph → Exploration lifecycle stages → Fieldcraft courses."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:F2")

r = 4
rows_readme = [
    ["Version", "1.0 — 2026-08-03"],
    ["Owner", "ArgantaEnergy subsurface / Fieldcraft faculty"],
    ["Purpose", "Every real basin, cycle, petroleum system, field, well and study artifact ArgantaEnergy has evidence for — as relational rows, not prose. Update here first; the Knowledge Base (Obsidian-style graph, Exploration tab, Fieldcraft labs) is generated FROM this workbook, never the reverse."],
    ["Confidentiality", "Contains ONLY public-domain / open-licence data (USGS public domain, Sodir NLOD-2.0, Equinor Open Data) and ArgantaEnergy's own shipped app content. A confidential 2025 Play-Based-Exploration joint-study report informed the GENERIC METHODOLOGY on the 'Report Sections' tab (CRS/CCRS logic, GCF formula, GRV two-surface method) — no project name, place name, well name, lead name or number from that report appears anywhere in this workbook."],
    ["How to update", "Add rows to the entity tabs (Basin, Basin Cycle, Field, Well, ...). Always fill Provenance + Source Citation ID. Add new sources to the Citations tab FIRST, then reference the citation_id — never paste a citation as free text in another tab."],
    ["Provenance ladder", "Every fact-bearing row states its dataNature: measured (instrument) · reported (published, not ours) · interpreted (geologist judgement) · derived (computed) · forecast (model output) · scenario (what-if / pre-drill exercise) · reference (industry doctrine / classification vocabulary). Missing data stays blank — never fabricated."],
    ["Citation ladder", "Cite the PARENT, not the compiler. Doust's book is cited only for what is genuinely his (the cycle-comparability thesis, his 6 grouping criteria, his 2 own papers) — see the Citations tab 'tier' column (P0 primary / P1 synthesis / P2 derived)."],
]
for label, val in rows_readme:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=True)
    c = ws.cell(row=r, column=2, value=val)
    c.font = BODY_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 15 + 13 * (len(val) // 90)
    r += 2

r += 1
ws.cell(row=r, column=1, value="Live counts (auto-updates as you add rows)").font = Font(name=FONT_NAME, size=12, bold=True)
LIVE_COUNTS_ROW = r + 1  # the "Live counts" formulas are written at the very end of this
# script (once every entity sheet exists and its exact last-data-row is known) — see
# the "README LIVE COUNTS" section at the bottom. Writing them here with a whole-column
# COUNTA(Sheet!A:A) would silently over-count: several sheets have an explanatory NOTE
# row a couple of rows below their table, which COUNTA would count as if it were a data
# row. Deferring to a bounded COUNTA(Sheet!A2:A{last_data_row}) range avoids that bug.

# ══════════════════════════════════════════════════════════════════════════════
# DATA DICTIONARY
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Data Dictionary")
dd_rows = [
    ["Citations", "citation_id", "Primary key, e.g. C-USGS-01", "Referenced by every other tab's *_citation_id column"],
    ["Citations", "tier", "P0 primary / P1 synthesis / P2 derived", "P0 = the work that first established the fact. P1 = a review/compilation organizing P0s. P2 = ArgantaEnergy's own computed output."],
    ["Citations", "verification_method", "adversarial web research (2 passes, 2026-08-02) / direct primary read / existing shipped app code / unverified-inference", "How ArgantaEnergy came to trust this citation — NOT the same as the citation itself being peer-reviewed"],
    ["Basin Cycle", "geodynamics", "pre-rift / extensional / sag / compressional", "ArgantaEnergy's own generic rift-basin vocabulary — NOT a specific cited external classification (see Basin-cycle framework concept)"],
    ["Basin Cycle", "units", "semicolon-separated formation names", "Cross-reference to 'Stratigraphic Units' tab rows sharing the same basin_cycle_id"],
    ["Opportunity", "maturity", "lead / prospect / drill-ready", "ONE entity, a state — never split into separate Lead/Prospect/Segment tabs (PBE + PRMS both treat this as maturation, not new objects)"],
    ["Opportunity", "gcf_*", "0.00–1.00 per component (source/timing-migration/reservoir/trap/seal)", "GCF = CoSg = product of the five components (Rose 2001 convention). gcf_total is the product, not summed."],
    ["Well / Wellbore", "—", "'Well' = surface slot/location; 'Wellbore' = each named penetration under it", "Volve's own wb data already separates these — e.g. well 'F-11' has 4 wellbores (F-11, F-11 A, F-11 B, F-11 T2) sharing one surface x/y"],
    ["Study Stages", "—", "mirrors apps/energy/src/tabs/exploration/registry.ts STUDY_STAGES verbatim", "Keep these two in sync by hand until an export script exists — this tab must never silently diverge from the shipped code"],
    ["Doust Figure Sourcing", "sourcing", "Own / External / Compiled", "Own = no source named, ask Doust directly. External = a named author/publisher holds rights, citation notwithstanding. Text-only map — no image bytes stored anywhere in this workbook"],
    ["Report Sections", "—", "a GENERIC Play-Based-Exploration report ToC", "Methodology only (CRS/CCRS, GCF, GRV, YTF) — no project-specific names, numbers or confidential facts of any kind"],
    ["ALL entity tabs", "provenance", "measured/reported/interpreted/derived/forecast/scenario/reference", "See README. A 'scenario' row (e.g. a pre-drill teaching case) must say so in its Notes column — never presented as a live decision"],
]
write_table(ws, ["Tab", "Column", "Values / format", "Notes"], dd_rows, col_widths=[20, 18, 46, 60])

# ══════════════════════════════════════════════════════════════════════════════
# CITATIONS — the master bibliography. Every other tab cites BY ID into this.
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Citations")
citations = [
    ["C-USGS-01", "P0", "USGS", "n.d.", "Province/TPS/AU boundaries dataset (metadata record)", "USGS ScienceBase", "item 60ad2fd7d34e4043c850edb3", "Public Domain", "Y", "adversarial web research (2026-08-02)", "province=descriptive spatial container, not process; may bundle multiple basins"],
    ["C-USGS-02", "P0", "USGS", "n.d.", "Companion province/TPS/AU dataset (metadata record)", "USGS ScienceBase", "item 60ad2fa1d34e4043c850ed98", "Public Domain", "Y", "adversarial web research (2026-08-02)", "corroborates C-USGS-01 independently"],
    ["C-USGS-03", "P0", "USGS", "2000", "World Petroleum Assessment 2000 — Petroleum Systems chapter", "USGS DDS-060", "PS.pdf", "Public Domain", "Y", "adversarial web research (2026-08-02)", "TPS formal definition; 8-digit ID scheme"],
    ["C-USGS-04", "P0", "USGS", "2000", "World Petroleum Assessment 2000 — Introduction chapter", "USGS DDS-060", "IN.pdf", "Public Domain", "Y", "adversarial web research (2026-08-02)", "937 provinces / 128 assessed / 159 TPS / 270 AU counts"],
    ["C-USGS-05", "P0", "USGS", "n.d.", "USGS national/world assessment data catalog record", "USGS data.usgs.gov", "60bfec21d34e86b938917fa7.xml", "Public Domain", "Y", "adversarial web research (2026-08-02)", "AU/province topology overlap; ~1:5,000,000 screening-scale geometry"],
    ["C-DOUST-01", "P1", "Doust, H.", "2026", "Dissecting Sedimentary Basins (booklet, 102pp, 49 figures)", "self-published PDF, VU Amsterdam / KNGMG", "founder-supplied Google Drive file; local working copy kept at docs/arganta-energy/knowledge-base/doust-basin-figures/_source/ (gitignored, never committed)", "UNRESOLVED — reference only", "n/a", "direct primary read, complete 102-page PDF (2026-08-03)", "cycle-as-comparable-unit thesis + 6 grouping criteria are HIS; do not reproduce text/figures. All 49 figures classified Own/External/Compiled on the 'Doust Figure Sourcing' tab"],
    ["C-DOUST-02", "P0", "Doust, H.", "2003", "Placing petroleum systems and plays in their basin history context", "First Break 21(9): 73-83", "", "assume publisher copyright — cite, do not reproduce", "n/a", "cited in Doust's own bibliography (direct read)", "Doust's own genuinely original paper"],
    ["C-DOUST-03", "P0", "Beglinger, S.E.; Corver, M.P.; Doust, H.; Cloetingh, S.; Thurmond, A.K.", "2012", "A new approach to relating petroleum system and play development to basin evolution", "AAPG Bulletin 96(6): 953-982", "", "assume publisher copyright — cite, do not reproduce", "n/a", "cited in Doust's own bibliography (direct read)", "Doust's own genuinely original paper"],
    ["C-KINGSTON-83", "P0", "Kingston, D.R. et al.", "1983", "Global basin classification system", "AAPG Bulletin", "", "assume publisher copyright", "N", "named in Doust's bibliography; NOT independently verified by ArgantaEnergy (2 research passes, 2026-08-02, 107 agents)", "do NOT cite specific claims about this scheme as confirmed fact"],
    ["C-SODIR-VOLVE", "P0", "Sodir (Norwegian Offshore Directorate)", "ongoing", "Volve field factpage, NPDID 3420717", "Sodir Factpages", "factpages.sodir.no", "NLOD-2.0", "Y", "existing shipped app code (src/atlas/volve.ts)", "field/well/company/licence identifiers"],
    ["C-EQUINOR-VOLVE", "P0", "Equinor", "2018", "Volve open dataset (Eclipse model, logs, production)", "Equinor Open Data", "", "Equinor Open Data licence", "Y", "existing shipped app code (public/wb/*)", "wells, logs, PVT, contacts, production"],
    ["C-KIEFT-MILTON", "P0", "Kieft, R.; Milton, N. et al.", "n.d.", "Volve Hugin Fm regional stratigraphy (peer reference)", "peer-reviewed (full citation not re-verified this session)", "", "assume publisher copyright — cite, do not reproduce", "N", "cited in shipped app code (explData.ts) — not independently re-verified this session", "unit ages/roles in Stratigraphic Units tab"],
    ["C-USGS-DDS69", "P0", "USGS", "2012", "World Assessment of Undiscovered Oil and Gas Resources (2012 update)", "USGS DDS-069", "apps/energy/public/world/*.geojson,*.json", "Public Domain", "Y", "existing shipped app code (apps/energy/scripts/extract-usgs-world.py)",
     "A DIFFERENT ASSESSMENT VINTAGE from C-USGS-01..05 (which characterize the 2000 baseline, DDS-60: 937 provinces / 128 assessed). This 2012 revision, as actually ingested here, has 179 provinces / 340 assessment units — not a contradiction, a different release. This is the citation for every bulk-populated Region/Country/Province/Petroleum-System/Assessment-Unit row in this workbook."],
    ["C-GOGET-01", "P0", "Global Energy Monitor (GEM)", "2026", "Global Oil and Gas Extraction Tracker (March 2026 release)", "Global Energy Monitor", "data-energy/raw/goget/Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx", "CC-BY-4.0 — ATTRIBUTION REQUIRED on redistribution", "Y", "existing shipped app code (src/atlas/goget.ts, scripts/build-osdu.mjs)",
     "8,032 real global fields. Citation for every GOGET-sourced Field row and every province the field-to-province spatial crosswalk (scripts/build-cockpit-spatial.mjs) assigned."],
    # ── primary sources named IN Doust's own figure captions/references — the actual
    # rightsholders for the 31 "External" rows on the Doust Figure Sourcing tab. Citing
    # them here is what the citation ladder means: ARE never reproduce Doust's figures
    # or theirs; we cite the parent for the underlying geological fact/example. ────────
    ["C-ZIEGLER-04", "P0", "Ziegler, P.A.; Cloetingh, S.", "2004", "Dynamic processes controlling evolution of rifted basins", "Earth Science Reviews 64(1-2): 1-50", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 2 (rift-triggering stresses)"],
    ["C-FRASER-07", "P0", "Fraser, S.I.; Fraser, A.J.; Lentini, M.R.; Gawthorpe, R.L. (eds)", "2007", "Return to rifts – the next wave", "Petroleum Geoscience 13(2): 99-104", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 3 (crustal extension models)"],
    ["C-KATZ-95", "P0", "Katz, B.J.", "1995", "A survey of rift basin source rocks (in Lambiase, J. ed., Hydrocarbon habitat in rift basins)", "Geol. Soc. Special Publ. 80: 213-242", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 15 (continental synrift stratigraphy)"],
    ["C-WIKI-DEPENV", "P1", "Wikipedia contributors", "n.d.", "Depositional environment", "en.wikipedia.org/wiki/Depositional_environment", "", "CC-BY-SA 4.0 — attribution required, different chain from the academic sources here", "Y", "named in Doust's own caption (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 6"],
    ["C-PERNER-18", "P0", "Perner, K. et al.", "2018", "Lower Rhine Graben Quaternary sediment thickness study", "Petroleum Geoscience 24(4): 425-439", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 13 (Lower Rhine Graben)"],
    ["C-JIANGSHU-13", "P0", "Jiang Shu et al.", "2013", "Sequence stratigraphic architecture and sand body distribution, lacustrine basins, east China", "AAPG Bull 97(9): 1447-1475", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 16 (Liaodong Bay)"],
    ["C-LUNDIN-18", "P0", "Lundin, E.R.; Dore, A.G.F.; Redfield, T.F.", "2018", "Magmatism and extension rates at rifted margins", "Petroleum Geoscience 24(4): 379-392", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 18 (magma-poor vs magma-rich margins)"],
    ["C-TAKANO-02", "P0", "Takano, O.", "2002", "Evolution in a rifted & inverted basin", "Sedimentary Geology 152: 79-97", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 19 (Niigata Basin)"],
    ["C-MIALLBOOK-19", "P1", "Miall, A.D. (ed)", "2019", "The Sedimentary Basins of the United States and Canada", "Elsevier, ISBN 978-0-444-63895-3", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "edited volume; sources Doust figs. 24 (ch.5 Miall), 36 (Taconic, ch. Ettensohn et al), 37 (Laramide, ch. Lawton)"],
    ["C-WATTS-DALY-18", "P0", "Watts, A.B. et al., in Daly, M.C.; Fuck, R.A.; Julia, J.; Macdonald, D.I.M. (eds)", "2018", "Cratonic basin formation studies (Watts et al. chapter, in the Parnaiba Basin cratonic-basins volume)", "Geol. Soc. Sp. Publ. 472", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust figs. 25, 28 (intracratonic basin comparisons, Taoudenni Basin)"],
    ["C-CRAIG-10", "P0", "Craig, J.; Grigo, D.; Rebora, A.; Serafini, G.; Tebaldi, E.", "2010", "Neoproterozoic to Early Cenozoic geology, N. Africa and Middle East", "Proc. 7th Petroleum Geology Conf., Geol. Soc. London: 673-705", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "co-sources Doust fig. 28 (Taoudenni Basin)"],
    ["C-BURGESS-19", "P0", "Burgess, P.A.", "2019", "Phanerozoic evolution, sedimentary cover of the N. American Platform", "in Miall (ed) Sedimentary basins of USA and Canada, Elsevier: 39-75", "doi.org/10.1016/B978-0-444-63985-3.00002-4", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 27 (Illinois Basin, column annotated w/ Sloss 1963 megacycles)"],
    ["C-ALLEN-91", "P0", "Allen, P.A.; Crampton, S.L.; Sinclair, H.D.", "1991", "The inception and early evolution of the North Alpine foreland basin", "Basin Research 3: 143-163", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 32 (Molasse Basin, Switzerland)"],
    ["C-DERUIG-06", "P0", "de Ruig, M.J.; Hubbard, S.M.", "2006", "Seismic facies and reservoir characteristics of a deep marine channel belt in the Molasse foreland basin, Puchkirchen Formation, Austria", "AAPG Bull 90(5): 735-752", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust figs. 33, 34 (Austrian Molasse Basin)"],
    ["C-MANN-06", "P0", "Mann, P.; Escalona, A.; Castillo, M.V.", "2006", "Regional geologic and tectonic setting of the Maracaibo supergiant basin, western Venezuela", "AAPG Bull 90(4): 445-477", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 35 (Maracaibo Basin)"],
    ["C-BUCHS-09", "P0", "Buchs, D.M. et al.", "2009", "Late Cretaceous to Miocene seamount accretion and melange formation, Osa and Burica peninsulas, southern Costa Rica", "journal citation truncated in source PDF", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 38 (Costa Rica forearc model)"],
    ["C-HENRY-12", "P0", "Henry, P.; Kanamatsu, T.; Moe, K.T.; Strasser, M. et al. (IODP Exp. 333 Scientific Party)", "2012", "IODP Expedition 333: Return to the Nankai Trough subduction inputs sites and coring of mass transport deposits", "Scientific Drilling 14: 4-17", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 39 (Nankai Trough)"],
    ["C-DEVILLE-03", "P0", "Deville, E.; Mascle, A.; Guerlais, S-H.; Decalf, C.; Colletta, B.", "2003", "Lateral changes in frontal accretion and mud-volcanism processes in the Barbados accretionary prism", "AAPG Memoir 79: 656-674", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 40 (Caribbean Island Arc)"],
    ["C-SPEED-91", "P0", "Speed, R.C.; Barker, L.H.; Payne, P.L.B.", "1991", "Geology and hydrocarbon evolution of Barbados", "Jl. Petroleum Geol. 14(3): 323-342", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 41 (Barbados outer ridge)"],
    ["C-SURYANUGRAHA-13", "P0", "Surya Nugraha, A.M.; Hall, R.", "2013", "Cenozoic stratigraphy of the east Java forearc", "Berita Sedimentologi 26(5): 5-17", "iagi.or.id", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 42 (Java forearc)"],
    ["C-TARAPOANCA-04", "P0", "Tarapoanca, M.", "2004", "Architecture, 3D geometry and tectonic evolution of the Carpathians foreland basin", "PhD thesis, Vrije Universiteit Amsterdam", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 43 (Romania foreland basin, seismic line drawing)"],
    ["C-MANCEDA-95", "P0", "Manceda, R.; Figueroa, D.", "1995", "Inversion of the Mesozoic Neuquen rift in the Malargue fold and thrust belt, Mendoza, Argentina", "AAPG Memoir 62: 369-382", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust fig. 44 (Neuquen Basin, Argentina)"],
    ["C-TEISSERENC-89", "P0", "Teisserenc, P.; Villemin, J.", "1989", "Sedimentary basin of Gabon – geology & oil systems", "AAPG Memoir 48: 117-200", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "co-sources Doust fig. 45 (Gabon Basin)"],
    ["C-PETRONAS-99", "P0", "Petronas", "1999", "The Petroleum Geology and Resources of Malaysia", "Petronas, ISBN 983-9738-10-0", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "co-sources Doust fig. 45 (Malay Basin)"],
    ["C-DOUST-SUMNER-07", "P0", "Doust, H.; Sumner, H.S.", "2007", "Petroleum systems in rift basins – a collective approach in Southeast Asian Tertiary basins", "Petroleum Geoscience 13(2): 127-144", "", "assume publisher copyright — cite, do not reproduce", "n/a", "cited in Doust's own bibliography (direct read)", "Doust's own genuinely original paper — sources fig. 46; a simpler permission case than the true third-party figures"],
    ["C-RUPPRECHT-18", "P0", "Rupprecht, B.J.; Sachsenhofer, R.F.; Zach, V.; Bechtel, A.; Gratzer, R.; Kucher, F.", "2018", "Oil and gas in the Vienna Basin: Hydrocarbon generation and alteration in a classical hydrocarbon province", "Petroleum Geoscience 25(1): 3-29", "", "assume publisher copyright — cite, do not reproduce", "N", "named in Doust's bibliography (direct read, verified against complete PDF 2026-08-03)", "sources Doust figs. 48, 49 (Vienna Basin)"],
]
r = write_table(ws, ["citation_id", "tier", "authors", "year", "title", "publication", "identifier", "licence_status", "verified", "verification_method", "notes"], citations,
                col_widths=[15, 6, 26, 8, 40, 22, 26, 24, 8, 34, 46])
LAST_ROW["Citations"] = r - 1
note_row(ws, r + 1, "No confidential material is cited here. A 2025 confidential Play-Based-Exploration joint-study report informed the GENERIC methodology described on 'Report Sections' only — it is never named or cited as a source in this workbook.", 11)

# ══════════════════════════════════════════════════════════════════════════════
# CONCEPTS / GLOSSARY — real, already-shipped (knowledge-model.ts CONCEPTS + new)
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Concepts")
concepts = [
    ["prms", "PRMS", "Petroleum Resources Management System — SPE reserves/resources classification.", "Classifies recoverable volumes by commercial maturity/uncertainty: Reserves (1P/2P/3P), Contingent Resources, Prospective Resources.", "classification", ""],
    ["volumetrics-stoiip", "Volumetrics (STOIIP)", "Stock-tank oil initially in place.", "STOIIP = GRV x NTG x phi x (1-Sw) / Boi. Recoverable = STOIIP x Recovery Factor.", "volumetrics", ""],
    ["recovery-factor", "Recovery Factor", "Fraction of in-place hydrocarbon ultimately produced.", "Depends on drive mechanism, rock quality, development. From analogs, material balance, simulation.", "volumetrics", ""],
    ["archie", "Archie", "Water saturation from resistivity.", "Sw = ((a x Rw) / (phi^m x Rt))^(1/n). a/m/n calibrated from core.", "petrophysics", ""],
    ["vrr", "VRR", "Voidage Replacement Ratio.", "VRR = sum(injection voidage)/sum(production voidage). ~1.0 balanced.", "reservoir-management", ""],
    ["dca", "Decline Curve Analysis", "Empirical production forecasting.", "Fits Arps declines (exponential/hyperbolic/harmonic) to rate-time history.", "reservoir-management", ""],
    ["fpot", "Reservoir Potential (FPOT)", "First-Point-of-Time unconstrained potential.", "Rate a well/field could produce with no surface constraint.", "reservoir-management", ""],
    ["gcos", "GCoS / GCF", "Geological Chance of Success (= Chance of Geological Success, CoSg).", "Product of independent risk factors (source, reservoir, seal, trap, timing). CoSg = CoS_play x CoS_prospect.", "risk", ""],
    ["stage-gate", "Stage-Gate", "Decision-gated project maturation.", "Projects advance through gates (G0...G3) with a fund/hold/kill decision at each.", "process", ""],
    ["history-match", "History Match", "Calibrating the dynamic model to observed data.", "Adjusts uncertain parameters until simulation reproduces measured pressure/production.", "reservoir-engineering", ""],
    ["waterflood", "Waterflood", "Secondary recovery by water injection.", "Maintains pressure, sweeps oil to producers. Managed via VRR balance.", "reservoir-engineering", ""],
    ["provenance", "Provenance", "The data-truth ladder.", "measured / interpreted / derived / reference (KB scheme) — every value states its class.", "doctrine", ""],
    ["evidence-native", "Evidence-native", "Every claim cites its source.", "No value without a source + provenance class. Missing data flagged, never faked.", "doctrine", ""],
    ["basin-cycle-framework", "Basin-cycle framework", "The classification approach behind every basin's cycle stack.", "Cycle (not whole basin) is the comparable unit across basins — see Citations C-DOUST-01/02/03.", "classification", "C-DOUST-01"],
    ["geodynamics-extensional", "Extensional (rift) cycle", "Basin cycle formed by active crustal extension and normal faulting.", "Half-graben or symmetric grabens; synrift fill ranges non-marine (lacustrine) to marine, underfilled to overfilled per accommodation-vs-supply balance. Real example: Viking Graben early-climax and late-synrift cycles (see Basin Cycle tab). Named worked examples in the Doust Figure Sourcing tab (figs 13, 15, 16) cite Perner 2018 (Rhine Graben), Katz 1995 (synrift source rocks), Jiang Shu 2013 (Liaodong Bay) — cite those primaries directly, never Doust's figure.", "classification", "C-DOUST-01"],
    ["geodynamics-sag", "Sag (postrift / cratonic) cycle", "Basin cycle formed by thermal-relaxation subsidence, no active faulting.", "Postrift sag overlies rift cycles along passive margins (thin, widening); cratonic sag sits on Precambrian shields with megasequence stacking (Sloss 1963 nomenclature). Real example: Viking Graben post-rift sag cycle (Shetland Gp-Nordland Gp, see Basin Cycle tab). Named worked examples cite Lundin 2018 and Takano 2002 (postrift margins), Burgess 2019 and Watts et al. in Daly 2018 (cratonic sag) — see Doust Figure Sourcing tab.", "classification", "C-DOUST-01"],
    ["geodynamics-compressional", "Compressional (foreland / forearc) cycle", "Basin cycle formed by plate convergence — fold-thrust foreland loading or subduction forearc accretion.", "Foreland: flexural depocenter migrates continentward ahead of the thrust front (Molasse-type). Forearc: sediment traps between a volcanic arc and an accretionary prism. Named worked examples cite Allen et al. 1991 and de Ruig & Hubbard 2006 (Molasse Basin), Buchs 2009 and Henry 2012 (forearc) — see Doust Figure Sourcing tab. Not yet represented by a real BasinCycle row in this workbook (Volve's Viking Graben never entered compression).", "classification", "C-DOUST-01"],
]
LAST_ROW["Concepts"] = write_table(ws, ["concept_id", "title", "one_line_definition", "body", "category", "citation_id"], concepts,
                                    col_widths=[22, 24, 40, 60, 20, 14]) - 1

# ══════════════════════════════════════════════════════════════════════════════
# DOUST FIGURE SOURCING — every one of the 49 figures in "Dissecting Sedimentary
# Basins", classified as Own (Doust's own drawing, no source named — ask HIM for
# permission) vs External (a specific named author/publisher holds the rights,
# citation notwithstanding) vs Compiled (multiple unnamed sources). Verified against
# the COMPLETE source PDF (102 pages) on 2026-08-03, closing an earlier gap where a
# partial text extraction stopped mid-page-79. No image bytes are stored anywhere in
# this workbook or the app — text-only sourcing map, feeding the permission
# conversation with Doust and the citation ladder for any future ArgantaEnergy-drawn
# (i.e. NOT reproduced) basin-type diagrams.
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Doust Figure Sourcing")
doust_figs = [
    [1, 9, "Typical sedimentary basins related to tectonic setting", "Ch1 foundational", "foundational", "External", "Kingston et al. 1983", "C-KINGSTON-83", "figure built to illustrate Kingston's classification criteria"],
    [2, 12, "Stresses that trigger stretching, basin development, crustal separation", "Ch1 foundational", "foundational", "External", "Ziegler & Cloetingh 2004", "C-ZIEGLER-04", ""],
    [3, 13, "Models illustrating crustal reaction to extensional stresses", "Ch1 foundational", "foundational", "External", "Fraser et al. 2007", "C-FRASER-07", ""],
    [4, 14, "Three mechanisms producing subsidence/accommodation space", "Ch1 foundational", "foundational", "Own", "", "", ""],
    [5, 14, "Examples of single and multiple cycle basins", "Ch1 foundational", "foundational", "Own", "", "", "caption itself uncited; a nearby unrelated line mentions Lahusen 1997 but reads as general-text attribution, not this caption's source — worth Doust confirming directly"],
    [6, 17, "General depositional environment model", "Ch1 foundational", "foundational", "External", "Wikipedia: Depositional environment", "C-WIKI-DEPENV", "CC-BY-SA licence chain, different from the academic sources"],
    [7, 22, "Sketch cross sections and map views of common rift geometries", "Ch2 Divergent", "extensional", "Own", "", "", ""],
    [8, 23, "Half-graben failed-rift cycle geometry, S. Viking Graben", "Ch2 Divergent", "extensional", "Own", "", "", "real basin (Viking Graben), no source named"],
    [9, 23, "Symmetrical failed-rift geometries: Viking Graben + Malay Basin", "Ch2 Divergent", "extensional", "Own", "", "", "real basins, no source named"],
    [10, 24, "Hyperextended rifts: Basin & Range NV, Laptev Sea", "Ch2 Divergent", "extensional", "Own", "", "", "Mohn et al 2011 cited nearby re: Alpine fold belt, not as this figure's source"],
    [11, 26, "Stages in development of a typical rift cycle", "Ch2 Divergent", "extensional", "Own", "", "", "Prosser 1993 / Changgui Xu et al 2025 discussed as competing schemes the cartoon synthesizes, not redrawn-from"],
    [12, 28, "Depositional environments/facies in climax stage of half-graben rift", "Ch2 Divergent", "extensional", "Own", "", "", ""],
    [13, 28, "Lower Rhine Graben, Germany — Quaternary sediment thickness", "Ch2 Divergent", "extensional", "External", "Perner et al. 2018", "C-PERNER-18", ""],
    [14, 30, "Impact of accommodation-space vs. sedimentation rate", "Ch2 Divergent", "extensional", "Own", "", "", ""],
    [15, 31, "Stratigraphy of continental synrift sequences, several continents", "Ch2 Divergent", "extensional", "External", "Katz 1995", "C-KATZ-95", ""],
    [16, 32, "Liaodong Bay, East China — Oligocene synrift depositional environments", "Ch2 Divergent", "extensional", "External", "Jiang Shu et al. 2013", "C-JIANGSHU-13", ""],
    [17, 41, "Orange Basin, South Africa — passive margin profile", "Ch3 Unfaulted postrift sag", "sag-postrift", "Own", "", "", ""],
    [18, 45, "Slowly- vs rapidly-extending margin comparison", "Ch3 Unfaulted postrift sag", "sag-postrift", "External", "Lundin et al. 2018", "C-LUNDIN-18", ""],
    [19, 46, "Niigata Basin, Japan — marginal rift/postrift backarc", "Ch3 Unfaulted postrift sag", "sag-postrift", "External", "Takano 2002", "C-TAKANO-02", ""],
    [20, 48, "West Atlantic passive margin cycle variety", "Ch3 Unfaulted postrift sag", "sag-postrift", "Compiled", "various sources (unnamed)", "", "caption says 'interpreted and compiled from various sources' — none individually named"],
    [21, 53, "Worldwide location of intracratonic basins on Precambrian cratons", "Ch4 Sag, continental interiors", "sag-cratonic", "Own", "", "", ""],
    [22, 55, "Arabian Basin cross section, adjacent to Arabian Shield", "Ch4 Sag, continental interiors", "sag-cratonic", "Own", "", "", ""],
    [23, 56, "Mesozoic carbonate shelf, Middle East (central Oman)", "Ch4 Sag, continental interiors", "sag-cratonic", "Own", "", "", ""],
    [24, 57, "Cambrian cratonic margin, western Canada", "Ch4 Sag, continental interiors", "sag-cratonic", "External", "Miall 2019 (ch.5, in Miall ed.)", "C-MIALLBOOK-19", ""],
    [25, 58, "Three intracratonic basins (S/N America, Africa) compared", "Ch4 Sag, continental interiors", "sag-cratonic", "External", "Watts et al. in Daly (eds) 2018", "C-WATTS-DALY-18", ""],
    [26, 59, "Williston Basin, USA — stacked megasequences", "Ch4 Sag, continental interiors", "sag-cratonic", "Own", "", "", ""],
    [27, 61, "Illinois Basin, USA — stratigraphic column + cross section", "Ch4 Sag, continental interiors", "sag-cratonic", "External", "Burgess 2019 (in Miall ed.); annotated w/ Sloss 1963 nomenclature", "C-BURGESS-19", ""],
    [28, 64, "Taoudenni Basin, Mali — Palaeozoic cratonic sag over Precambrian", "Ch4 Sag, continental interiors", "sag-cratonic", "External", "Watts et al. in Daly (eds) 2018 + Craig et al. 2010", "C-WATTS-DALY-18; C-CRAIG-10", ""],
    [29, 67, "Section across a typical foreland basin", "Ch5 Convergent", "compressional-foreland", "Own", "", "", ""],
    [30, 69, "Continent-continent collision, comparable-density plates", "Ch5 Convergent", "compressional-foreland", "Own", "", "", ""],
    [31, 70, "Ocean-continent collision, unequal-density plates", "Ch5 Convergent", "compressional-foreland", "Own", "", "", ""],
    [32, 73, "Molasse Basin, eastern Switzerland — foreland stratigraphy", "Ch5 Convergent", "compressional-foreland", "External", "Allen et al. 1991", "C-ALLEN-91", ""],
    [33, 73, "Austrian Molasse Basin structure", "Ch5 Convergent", "compressional-foreland", "External", "de Ruig & Hubbard 2006", "C-DERUIG-06", ""],
    [34, 74, "Oligocene palaeogeography, Molasse Basin near Salzburg", "Ch5 Convergent", "compressional-foreland", "External", "de Ruig & Hubbard 2006", "C-DERUIG-06", ""],
    [35, 75, "Maracaibo Basin, western Venezuela — sketch map", "Ch5 Convergent", "compressional-foreland", "External", "Mann et al. 2006", "C-MANN-06", ""],
    [36, 76, "Taconic foreland basin cycle (Upper Ordovician)", "Ch5 Convergent", "compressional-foreland", "External", "Miall ed. 2019 (Sedimentary Basins of US+Canada)", "C-MIALLBOOK-19", ""],
    [37, 79, "Laramide foreland basin cycle, western USA", "Ch5 Convergent", "compressional-foreland", "External", "Miall ed. 2019 (Sedimentary Basins of US+Canada)", "C-MIALLBOOK-19", ""],
    [38, 81, "Forearc zone model, Eocene of Costa Rica", "Ch5 Convergent (forearc)", "compressional-forearc", "External", "Buchs et al. 2009", "C-BUCHS-09", ""],
    [39, 82, "Nankai Trough forearc, SE Japan — seismic-based section", "Ch5 Convergent (forearc)", "compressional-forearc", "External", "Henry et al. 2012", "C-HENRY-12", ""],
    [40, 83, "Caribbean Island Arc — Tobago/Barbados forearc basins", "Ch5 Convergent (forearc)", "compressional-forearc", "External", "Deville et al. 2003", "C-DEVILLE-03", ""],
    [41, 84, "Barbados outer ridge and accretionary prism", "Ch5 Convergent (forearc)", "compressional-forearc", "External", "Speed et al. 1991", "C-SPEED-91", ""],
    [42, 86, "Java forearc Tertiary stratigraphy", "Ch5 Convergent (forearc)", "compressional-forearc", "External", "Surya Nugraha & Hall 2013", "C-SURYANUGRAHA-13", ""],
    [43, 92, "Seismic profile line drawing, foreland basin, Romania", "Ch6 Building the basin", "synthesis", "External", "Tarapoanca 2004", "C-TARAPOANCA-04", ""],
    [44, 93, "Fold belt cross section + palinspastic reconstruction, Neuquen Basin, Argentina", "Ch6 Building the basin", "synthesis", "External", "Manceda & Figueroa 1995", "C-MANCEDA-95", ""],
    [45, 94, "Synrift fills compared: Gabon Basin vs Malay Basin", "Ch6 Building the basin", "synthesis", "External", "Teisserenc & Villemin 1989 + Petronas 1999", "C-TEISSERENC-89; C-PETRONAS-99", ""],
    [46, 95, "Evolution cartoons, SE Asian Tertiary basins", "Ch6 Building the basin", "synthesis", "External", "Doust & Sumner 2007", "C-DOUST-SUMNER-07", "Doust's OWN prior co-authored paper — a simpler permission case than the true third-party figures"],
    [47, 96, "Basin trajectory plot, Almada-Camamu Basin, Brazil", "Ch6 Building the basin", "synthesis", "External", "Beglinger, Corver, Doust, Cloetingh & Thurmond 2012", "C-DOUST-03", "Doust's OWN prior co-authored paper — a simpler permission case than the true third-party figures"],
    [48, 98, "Vienna Basin cross section, strike-slip pull-apart", "Ch6 Building the basin", "synthesis", "External", "Rupprecht et al. 2018", "C-RUPPRECHT-18", ""],
    [49, 99, "Vienna Basin geological evolution, complex", "Ch6 Building the basin", "synthesis", "External", "Rupprecht et al. 2018", "C-RUPPRECHT-18", ""],
]
# image_file / permission close the loop to the app: the same fig-NN.png the
# Knowledge Bank's basin atlas renders, and the permission gate that governs it.
# Both the PNGs and the source PDF are gitignored — local-only until cleared.
for row in doust_figs:
    row.append(f"doust-figures/fig-{row[0]:02d}.png")
    # 'internal' = cleared by the organization (2026-08-03) for internal scientific /
    # educational use, CONDITIONAL ON ATTRIBUTION being displayed. Not 'public':
    # public redistribution would still need each rightsholder's own consent.
    row.append("internal")
LAST_ROW["Doust Figure Sourcing"] = write_table(ws, ["fig_no", "page", "caption", "chapter", "basin_type", "sourcing", "source_short", "citation_id", "notes",
                                                     "image_file", "permission"], doust_figs,
                col_widths=[8, 8, 46, 24, 20, 12, 40, 20, 60, 30, 12]) - 1
note_row(ws, LAST_ROW["Doust Figure Sourcing"] + 2,
         "17 Own (no source named — ask Doust directly): figs 4,5,7,8,9,10,11,12,14,17,21,22,23,26,29,30,31. 31 External (named source; permission runs through that author/publisher, not Doust): the rest except fig 20. 1 Compiled from unnamed sources: fig 20. Figs 46-47 are a special case: their named source is Doust's OWN earlier co-authored work. "
         "image_file points at apps/energy/public/doust-figures/ — extracted locally by extract_doust_figures.py and rendered by the Exploration Knowledge Bank's basin atlas. Those PNGs and the source PDF are GITIGNORED and must not be committed, deployed or shown to students while permission is 'pending'. Full narrative: docs/arganta-energy/knowledge-base/doust-basin-figures/README.md.", 11)

# ══════════════════════════════════════════════════════════════════════════════
# ID SCHEME — matches src/atlas/spine.ts's makeId(entity, authority, nativeId) =
# "atlas:{entity}:{authority}:{nativeId}" EXACTLY. Applied uniformly across every
# tab from here on, so a future importer can match rows against the real ATLAS
# spine with zero remapping. Constants for the Volve thread use the SAME values
# already shipped in src/atlas/volve.ts wherever a real one exists (field, company,
# licence, asset, reservoir) — replacing the earlier hand-typed shortcuts.
# ══════════════════════════════════════════════════════════════════════════════
def aid(entity, authority, native):
    return f"atlas:{entity}:{authority}:{native}"


REGION_EUROPE = aid("region", "usgs", "4")
COUNTRY_NORWAY = aid("country", "un", "NO")           # matches makeId('country','un','NO') in atlas/volve.ts
PROVINCE_4025 = aid("province", "usgs", "4025")
BASIN_VIKING = aid("basin", "atlas", "viking-graben")  # ArgantaEnergy's own geodynamic refinement, not a USGS-native id
TPS_KIMMERIDGIAN = aid("petroleum-system", "usgs", "402501")
AU_VIKING = aid("assessment-unit", "usgs", "40250101")
PLAY_HUGIN = aid("play", "atlas", "hugin")
OPP_VOLVE = aid("opportunity", "atlas", "volve-predrill")
FIELD_VOLVE = aid("field", "sodir", "3420717")         # matches makeId('field','sodir','3420717') in atlas/volve.ts
RESERVOIR_HUGIN = aid("reservoir", "atlas", "volve-hugin")  # matches makeId('reservoir','atlas','volve-hugin')
COMPANY_EQUINOR = aid("company", "sodir", "32011216")  # matches makeId('company','sodir','32011216')
LICENCE_046BS = aid("licence", "sodir", "046BS")       # matches makeId('licence','sodir','046BS')
ASSET_VOLVE = aid("asset", "atlas", "volve")           # matches makeId('asset','atlas','volve')


def cycle_id(suffix):
    return aid("basin-cycle", "atlas", f"viking-graben-{suffix}")


def wellbore_id(name):
    return aid("wellbore", "sodir", slug(name))


def well_id(name):
    return aid("well", "sodir", slug(name))


# ══════════════════════════════════════════════════════════════════════════════
# WHERE axis: Region, Country, Province — WORLD-SCALE, real USGS 2012 (DDS-69)
# data already ingested in apps/energy/public/world/*. Every row here is real;
# 4025/Europe/Norway are the SAME real entities as the earlier hand-built rows,
# now sourced from the authoritative bulk file instead of hand-typed (verified
# 2026-08-03 to match exactly: AU 40250101 "Viking Graben", TPS "Kimmeridgian
# Shales", prvCode 4025 "North Sea Graben", regCode 4 "Europe").
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Region")
region_rows = [[aid("region", "usgs", r["code"]), r["code"], r["name"], "reference", "C-USGS-DDS69", ""] for r in WORLD["regions"]]
LAST_ROW["Region"] = write_table_fast(ws, ["region_id", "code", "name", "provenance", "source_citation_id", "notes"], region_rows,
            col_widths=[16, 8, 30, 14, 18, 30]) - 1

ws = add_sheet("Country")
# Norway keeps the pre-established un:NO id (matches atlas/volve.ts); all others use
# a goget-authority id since GOGET's own Country/Area label is their only real key here.
country_rows = []
for c in WORLD["countries"]:
    cid = COUNTRY_NORWAY if c["name"] == "Norway" else aid("country", "goget", slug(c["name"]))
    country_rows.append([cid, c["name"], "reference" if c["hasAggregate"] else "reported",
                          "C-USGS-DDS69" if c["hasAggregate"] else "C-GOGET-01",
                          c["oilMean"], c["gasMean"], c["boeMean"]])
LAST_ROW["Country"] = write_table_fast(ws, ["country_id", "name", "provenance", "source_citation_id", "oilMean_mmbbl", "gasMean_bcf", "boeMean_mmboe"],
            country_rows, col_widths=[26, 22, 14, 18, 14, 14, 14]) - 1
note_row(ws, LAST_ROW["Country"] + 3, f"{len(WORLD['countries'])} countries = every distinct 'Country/Area' string across the real 8,032 GOGET fields (C-GOGET-01). oilMean/gasMean/boeMean are USGS DDS-69 country-level resource aggregates (C-USGS-DDS69) joined by exact name match where available — many GOGET countries have no USGS aggregate (no shared assessed province), left blank, not fabricated.", 7)

ws = add_sheet("Province")
province_rows = [
    [aid("province", "usgs", p["prvCode"]), p["prvCode"], p["prvName"], aid("region", "usgs", p["regCode"]) if p["regCode"] else "",
     "Y" if p["prvCode"] in {ps["prvCode"] for ps in WORLD["petroleumSystems"]} else "N",
     p["oilMean"], p["gasMean"], p["boeMean"],
     "~1:5,000,000 scale, 500 m fuzzy tolerance, no datum/spheroid defined — screening scale only",
     "reference", "C-USGS-DDS69", "may bundle multiple genetically-related basins — a province is a container, not itself a basin"]
    for p in WORLD["provinces"]
]
LAST_ROW["Province"] = write_table_fast(ws, ["province_id", "code", "name", "region_id", "assessed", "oilMean_mmbbl", "gasMean_bcf", "boeMean_mmboe",
                  "geometry_caveat", "provenance", "source_citation_id", "notes"], province_rows,
            col_widths=[16, 8, 26, 16, 8, 12, 12, 12, 44, 14, 18, 46]) - 1
note_row(ws, LAST_ROW["Province"] + 2,
         f"{len(WORLD['provinces'])} provinces = the FULL USGS 2012 World Assessment (DDS-69) province set, as ingested (C-USGS-DDS69) — a different assessment VINTAGE from the 2000 baseline (DDS-60) that the independently-verified USGS research (C-USGS-01..05) characterized as 937 provinces / 128 assessed. These are not contradictory counts — different USGS releases. 'assessed' here = has >=1 petroleum system in THIS revision's data, i.e. present in the Petroleum System tab.", 12)

ws = add_sheet("Block-Licence")
write_table(ws, ["block_id", "name", "country_id", "province_id", "operator", "status", "provenance", "source_citation_id"], [],
            col_widths=[16, 20, 14, 14, 20, 14, 14, 18])
note_row(ws, 2, "DEFERRED — no block/licence-level acreage data ingested yet (product decision 2026-08-02: hide until real data exists). Columns are fixed so this tab can be populated later with no schema change.", 8)

# ══════════════════════════════════════════════════════════════════════════════
# GEOLOGY axis: Basin (179 — 178 generic 1:1-from-province + 1 refined: Viking
# Graben), Basin Cycle, Stratigraphic Units, Petroleum System (211, world-scale),
# PS x Cycle join, Assessment Unit (340, world-scale), Play, Opportunity
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Basin")
basin_rows = []
for p in WORLD["provinces"]:
    prov_id = aid("province", "usgs", p["prvCode"])
    if p["prvCode"] == "4025":
        basin_rows.append([BASIN_VIKING, "Viking Graben", "failed-rift graben, Norwegian North Sea", prov_id, "Y",
            "The ATLAS spine (src/atlas/) currently types the province (4025) itself as 'basin', and names the AU (40250101) \"Viking Graben\" — a shorthand from earlier work. This row is the genuinely geodynamic Basin entity the verified USGS research shows is missing (province=container, not a basin). Reconciling ATLAS's tier names is a separate, not-yet-done refactor.",
            "interpreted", "C-USGS-01"])
    else:
        native_names = WORLD["provinceBasinNames"].get(p["prvCode"], [])
        note = ("1:1 seeded from its province name — a known simplification (a province may host several distinct basins); "
                + (f"GOGET's own field-level 'Basin' reports for this province: {'; '.join(native_names[:5])}"
                   + (f" (+{len(native_names)-5} more)" if len(native_names) > 5 else "")
                   if native_names else "no GOGET field-level basin names on record for this province"))
        basin_rows.append([aid("basin", "usgs", p["prvCode"]), p["prvName"], "", prov_id, "N", note, "derived", "C-USGS-DDS69"])
LAST_ROW["Basin"] = write_table_fast(ws, ["basin_id", "name", "setting", "province_id", "naming_collision_flag", "naming_collision_note",
                  "provenance", "source_citation_id"], basin_rows, col_widths=[26, 26, 26, 16, 10, 70, 12, 18]) - 1
note_row(ws, LAST_ROW["Basin"] + 2, f"{len(basin_rows)} basins = 1 refined (Viking Graben, real BasinCycle children below) + {len(basin_rows)-1} generic 1:1-from-province seeds. The 1:1 seeds are DERIVED, not independently researched — refine any of them the same way Viking Graben was refined, by adding real BasinCycle rows.", 8)

ws = add_sheet("Basin Cycle")
# ageMa = [top, base] Ma — oldest -> youngest order for readability
cycle_rows = [
    [cycle_id("pre-rift"), "Pre-rift basin fill (Triassic)", BASIN_VIKING, 237, 201, "pre-rift", "pre-rift / early basin fill",
     "non-marine", "", "continental", "clastic (fluvial redbeds)", "N", "fluvial", "unconformity", "reservoir",
     "Skagerrak Fm", "interpreted", "C-KIEFT-MILTON"],
    [cycle_id("early-climax-synrift"), "Early-climax syn-rift (Middle Jurassic)", BASIN_VIKING, 170, 157, "extensional", "early-climax syn-rift",
     "mixed", "proximal", "temperate", "clastic (sandstone)", "Y", "fluvial to shallow-marine", "transition", "reservoir",
     "Sleipner Fm; Hugin Fm", "interpreted", "C-KIEFT-MILTON"],
    [cycle_id("late-synrift"), "Late syn-rift (Late Jurassic)", BASIN_VIKING, 168, 145, "extensional", "late syn-rift",
     "marine", "distal", "temperate", "mudstone / shale", "N", "offshore to anoxic marine", "unconformity (BCU)", "source",
     "Heather Fm; Draupne Fm", "interpreted", "C-KIEFT-MILTON"],
    [cycle_id("postrift-sag"), "Post-rift sag (Cretaceous-Recent)", BASIN_VIKING, 100, 0, "sag", "post-rift sag",
     "marine", "distal", "temperate to boreal (cooling)", "mixed (mudstone / chalk / sandstone)", "Y", "marl/chalk to shallow-marine to glaciomarine", "transition", "mixed (seal-dominant)",
     "Shetland Gp; Ty Fm; Hordaland Gp; Utsira Fm; Nordland Gp", "interpreted", "C-KIEFT-MILTON"],
]
r = write_table(ws, ["cycle_id", "title", "basin_id", "age_top_ma", "age_base_ma", "geodynamics", "stage", "fill",
                      "proximity", "climate", "lithology", "environment_changed", "facies_associations", "boundary_type",
                      "dominant_role", "units", "provenance", "source_citation_id"], cycle_rows,
                nature_col=16, col_widths=[40, 30, 26, 8, 8, 12, 20, 8, 10, 16, 24, 10, 22, 16, 14, 40, 14, 18])
LAST_ROW["Basin Cycle"] = r - 1
note_row(ws, r + 1, "geodynamics/stage labels are ArgantaEnergy's own generic rift-basin vocabulary (see 'basin-cycle-framework' concept) — NOT a cited external classification. Cycle-comparability METHOD follows Doust (C-DOUST-01/02/03); the specific label set does not.", 18)

ws = add_sheet("Stratigraphic Units")
# name, group, ageMa[top,base], env, ps_role, role_note, basin_cycle_id, nature
strat_rows = [
    ["Nordland Gp", "Nordland", 23, 0, "marine to glaciomarine", "overburden", "", cycle_id("postrift-sag"), "interpreted", "C-KIEFT-MILTON"],
    ["Utsira Fm", "Nordland", 15, 3, "shallow-marine sand", "overburden", "regional aquifer", cycle_id("postrift-sag"), "interpreted", "C-KIEFT-MILTON"],
    ["Hordaland Gp", "Hordaland", 34, 15, "marine mudstone", "seal", "thick overburden seal", cycle_id("postrift-sag"), "interpreted", "C-KIEFT-MILTON"],
    ["Shetland Gp", "Shetland", 100, 56, "marl / chalk", "overburden", "", cycle_id("postrift-sag"), "interpreted", "C-KIEFT-MILTON"],
    ["Ty Fm", "Rogaland", 61, 58, "submarine fan sst", "reservoir", "secondary reservoir", cycle_id("postrift-sag"), "interpreted", "C-KIEFT-MILTON"],
    ["BCU", "-", 145, 145, "unconformity", "seal", "Base Cretaceous Unconformity - regional top-seal marker", cycle_id("late-synrift"), "interpreted", "C-KIEFT-MILTON"],
    ["Draupne Fm", "Viking", 157, 145, "anoxic marine shale", "source", "PRIMARY source + top seal (North Sea 'hot shale')", cycle_id("late-synrift"), "interpreted", "C-KIEFT-MILTON"],
    ["Heather Fm", "Viking", 168, 150, "offshore shale", "source", "secondary source + seal", cycle_id("late-synrift"), "interpreted", "C-KIEFT-MILTON"],
    ["Hugin Fm", "Vestland", 168, 157, "shallow-marine sst", "reservoir", "PRIMARY reservoir (diachronous, younging S)", cycle_id("early-climax-synrift"), "interpreted", "C-KIEFT-MILTON"],
    ["Sleipner Fm", "Vestland", 170, 165, "fluvial", "none", "", cycle_id("early-climax-synrift"), "interpreted", "C-KIEFT-MILTON"],
    ["Skagerrak Fm", "Hegre", 237, 201, "fluvial redbeds", "reservoir", "secondary reservoir", cycle_id("pre-rift"), "interpreted", "C-KIEFT-MILTON"],
]
LAST_ROW["Stratigraphic Units"] = write_table(ws, ["unit_name", "group", "age_top_ma", "age_base_ma", "environment", "ps_role", "role_note", "cycle_id",
                  "provenance", "source_citation_id"], strat_rows, nature_col=8, col_widths=[16, 12, 8, 8, 20, 12, 40, 40, 14, 18]) - 1

ws = add_sheet("Petroleum System")
ps_rows = []
for ps in WORLD["petroleumSystems"]:
    tps_id = aid("petroleum-system", "usgs", ps["tpsCode"])
    prov_id = aid("province", "usgs", ps["prvCode"])
    name = ps["tps"] or f"Unnamed (TPS {ps['tpsCode']})"
    if ps["tpsCode"] == "402501":
        ps_rows.append([tps_id, ps["tpsCode"], "Kimmeridgian Shales", prov_id, "Draupne Fm (Kimmeridge Clay equiv.)",
            "source + reservoir + seal + overburden rock, all petroleum genetically related to one pod (or related pods) of active source rock",
            "generation-migration-accumulation from the Draupne kitchen into the Hugin reservoir trap", "reference", "C-USGS-03"])
    else:
        ps_rows.append([tps_id, ps["tpsCode"], name, prov_id, "", "", "", "reference", "C-USGS-DDS69"])
LAST_ROW["Petroleum System"] = write_table_fast(ws, ["tps_id", "code", "name", "province_id", "source_rock_formation", "essential_elements_note",
                  "generation_migration_note", "provenance", "source_citation_id"], ps_rows,
            col_widths=[26, 10, 30, 16, 30, 50, 46, 14, 18]) - 1
note_row(ws, LAST_ROW["Petroleum System"] + 2,
         f"{len(ps_rows)} petroleum systems, world-scale (C-USGS-DDS69). tps_id's numeric code is DERIVED from the real USGS 8-digit AU-code structure (province digits + TPS digits, verified 2026-08-02) — not a synthetic ArgantaEnergy id. Only the Kimmeridgian Shales row (402501) has essential-elements/charge detail — everything else is name + province only, honestly blank rather than fabricated.", 9)

ws = add_sheet("PS x Cycle")
write_table(ws, ["tps_id", "cycle_id", "role_in_cycle", "notes"],
            [[TPS_KIMMERIDGIAN, cycle_id("late-synrift"), "source", "Draupne Fm source rock is deposited within this cycle"],
             [TPS_KIMMERIDGIAN, cycle_id("early-climax-synrift"), "reservoir", "Hugin Fm reservoir charged by this system is deposited within this cycle"]],
            col_widths=[26, 40, 14, 60])
note_row(ws, 4, "Explicit N:M join — a Petroleum System SPANS basin cycles (source in one, reservoir in another); it is never a single FK on either tab.", 4)

ws = add_sheet("Assessment Unit")
au_rows = [[aid("assessment-unit", "usgs", a["auCode"]), a["auCode"], a["auName"] or f"Unnamed AU {a['auCode']}",
            aid("petroleum-system", "usgs", a["prvCode"] + a["auCode"][4:6]), "Assessed", a["oilMean"], a["gasMean"], a["boeMean"],
            "reference", "C-USGS-DDS69"] for a in WORLD["assessmentUnits"]]
LAST_ROW["Assessment Unit"] = write_table_fast(ws, ["au_id", "code", "name", "tps_id", "status", "oilMean_mmbbl", "gasMean_bcf", "boeMean_mmboe",
                  "provenance", "source_citation_id"], au_rows, col_widths=[24, 10, 30, 26, 12, 12, 12, 12, 14, 18]) - 1

ws = add_sheet("Play")
LAST_ROW["Play"] = write_table(ws, ["play_id", "name", "cycle_id", "tps_id", "play_type", "description", "crs_status", "provenance", "source_citation_id"],
            [[PLAY_HUGIN, "Middle Jurassic Hugin", cycle_id("early-climax-synrift"), TPS_KIMMERIDGIAN,
              "structural (fault-bounded four-way dip closure)", "Middle Jurassic shallow-marine sandstone reservoir play, charged from the Draupne kitchen",
              "not started", "interpreted", "C-SODIR-VOLVE"]],
            nature_col=7, col_widths=[20, 20, 40, 26, 30, 46, 12, 14, 18]) - 1

ws = add_sheet("Opportunity")
opp_rows = [
    [OPP_VOLVE, "Volve (15/9-19 prospect) — pre-drill teaching case", PLAY_HUGIN, "drill-ready", "", "", "",
     0.90, 0.85, 0.72, 0.78, 0.82, round(0.90 * 0.85 * 0.72 * 0.78 * 0.82, 4),
     "oil", "", "", f"realized as field {FIELD_VOLVE} — 15/9-19 SR discovery, 1993",
     "scenario", "C-SODIR-VOLVE",
     "ILLUSTRATIVE, not a vetted geological chance assessment. These GCF component values are ArgantaEnergy's own placeholder defaults for the app's editable Exploration sliders (apps/energy/src/tabs/exploration/explData.ts), chosen to be reasonable given known Volve facts — NOT independently derived from mapped CRS evidence. Treat as a worked pedagogical example only, scored retrospectively against the real 1993 discovery."],
]
r = write_table(ws, ["opportunity_id", "name", "play_id", "maturity", "closure_area_km2", "closure_height_m", "lcc_tvdss_m",
                      "gcf_reservoir", "gcf_timing_migration", "gcf_source", "gcf_trap", "gcf_seal", "gcf_total",
                      "resource_type", "resource_p50", "resource_unit", "realized_as", "provenance", "source_citation_id", "notes"],
                opp_rows, nature_col=17, col_widths=[16, 32, 12, 12, 12, 12, 12, 10, 12, 10, 10, 10, 10, 10, 12, 12, 30, 14, 18, 70])
LAST_ROW["Opportunity"] = r - 1
note_row(ws, r + 1, "GCF component order here follows Rose 2001 convention (source, timing/migration, reservoir, trap, seal); gcf_total = PRODUCT of the five, not a sum. No confidential external prospect data appears in this tab.", 20)

# ══════════════════════════════════════════════════════════════════════════════
# ACCUMULATION: Field, Reservoir, Pool(deferred)
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Field")
field_rows = [
    [FIELD_VOLVE, "Volve", OPP_VOLVE, BASIN_VIKING, COUNTRY_NORWAY, "Equinor Energy AS", 1993, "15/9-19 SR",
     "Shut down", "OIL", 58.442, 1.888, "ED50 / UTM 31N", "TVDSS (m)", "measured", "C-SODIR-VOLVE"],
]
_country_id_by_name = {c["name"]: (COUNTRY_NORWAY if c["name"] == "Norway" else aid("country", "goget", slug(c["name"]))) for c in WORLD["countries"]}
for f in WORLD["fields"]:
    basin_id_val = aid("basin", "usgs", f["prvCode"]) if f["prvCode"] and f["prvCode"] != "4025" else (BASIN_VIKING if f["prvCode"] == "4025" else "")
    field_rows.append([
        aid("field", "goget", f["unitId"].lower()), f["name"], "", basin_id_val,
        _country_id_by_name.get(f["country"], ""), f["operator"] or "", f["discoveryYear"] or "",
        "", f["status"] or "", f["fuelType"] or "", f["lat"], f["lon"], "WGS84", "", "reported", "C-GOGET-01",
    ])
LAST_ROW["Field"] = write_table_fast(ws, ["field_id", "name", "opportunity_id", "basin_id", "country_id", "operator", "discovery_year",
                  "discovery_well", "status", "hc_type", "lat", "lon", "crs", "datum", "provenance", "source_citation_id"],
            field_rows, col_widths=[26, 34, 14, 22, 22, 22, 14, 14, 14, 16, 10, 10, 10, 10, 14, 18]) - 1
note_row(ws, LAST_ROW["Field"] + 2,
         f"{len(field_rows)} fields = 1 Volve (real Sodir/Equinor id, measured, deep data below) + {len(WORLD['fields'])} real GOGET fields (C-GOGET-01, GEM Global Oil and Gas Extraction Tracker, CC-BY-4.0 — attribution required on any redistribution). "
         f"basin_id resolved via a real spatial point-in-polygon join (fields against USGS province polygons): {WORLD['meta']['counts']['fieldsWithProvince']} of {len(WORLD['fields'])} GOGET fields ({100*WORLD['meta']['counts']['fieldsWithProvince']//len(WORLD['fields'])}%) matched a province — the rest are honestly blank (offshore/onshore areas outside any assessed province polygon in this revision), never guessed.", 16)

ws = add_sheet("Reservoir")
LAST_ROW["Reservoir"] = write_table(ws, ["reservoir_id", "field_id", "formation_name", "age", "lithology", "drive_mechanism",
                  "owc_tvdss_m", "contact_provenance_note", "provenance", "source_citation_id"],
            [[RESERVOIR_HUGIN, FIELD_VOLVE, "Hugin Fm", "Middle Jurassic", "shallow-marine sandstone", "waterflood",
              3200, "interpreted — DECK (EQUIL main structure)", "interpreted", "C-EQUINOR-VOLVE"]],
            nature_col=8, col_widths=[26, 22, 16, 16, 24, 14, 12, 32, 14, 18]) - 1

ws = add_sheet("Pool")
write_table(ws, ["pool_id", "reservoir_id", "field_id", "notes", "provenance", "source_citation_id"], [],
            col_widths=[14, 14, 10, 40, 14, 18])
note_row(ws, 2, "DEFERRED — Volve's data does not distinguish pools below reservoir level. Columns fixed for future use (e.g. a compartmentalized field).", 6)

# ══════════════════════════════════════════════════════════════════════════════
# WELLS: Well (surface slot) + Wellbore (each named penetration) — real wb data
# ══════════════════════════════════════════════════════════════════════════════
wb_wells = [
    {"name": "19 A", "well": "19", "x": 437506.858, "y": 6477887.468, "td_md": 4131, "td_tvd": 3318.72, "kb": "25.00m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": True},
    {"name": "19 BT2", "well": "19", "x": 437506.858, "y": 6477887.468, "td_md": 4250, "td_tvd": 3360.59, "kb": "25.00m", "role": "none", "has": {"logs": False, "traj": False, "production": False, "picks": True}, "is_exploration": True},
    {"name": "19 SR", "well": "19", "x": 437506.858, "y": 6477887.468, "td_md": 4644, "td_tvd": 3135.33, "kb": "25.00m", "role": "none", "has": {"logs": False, "traj": False, "production": False, "picks": True}, "is_exploration": True},
    {"name": "F-10", "well": "F-10", "x": 435052.248, "y": 6478559.613, "td_md": 5331, "td_tvd": 3017.01, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-11 A", "well": "F-11", "x": 435049.095, "y": 6478568.172, "td_md": 3762, "td_tvd": 3126.49, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-11 B", "well": "F-11", "x": 435049.095, "y": 6478568.172, "td_md": 4770, "td_tvd": 3257.3, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-11 T2", "well": "F-11", "x": 435049.095, "y": 6478568.172, "td_md": 4562, "td_tvd": 3400.3, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-11", "well": "F-11", "x": 435049.095, "y": 6478568.172, "td_md": 347, "td_tvd": 346.99, "kb": "54.90m", "role": "producer", "has": {"logs": False, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-12", "well": "F-12", "x": 435050.21, "y": 6478566.22, "td_md": 3520, "td_tvd": 3108.36, "kb": "54.90m", "role": "producer", "has": {"logs": True, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-14", "well": "F-14", "x": 435052.438, "y": 6478562.31, "td_md": 3750, "td_tvd": 3158.65, "kb": "54.90m", "role": "producer", "has": {"logs": True, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-15 A", "well": "F-15", "x": 435053.552, "y": 6478560.355, "td_md": 4095, "td_tvd": 3212.14, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-15 B", "well": "F-15", "x": 435053.552, "y": 6478560.355, "td_md": 3497, "td_tvd": 3016.5, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-15 C", "well": "F-15", "x": 435053.552, "y": 6478560.355, "td_md": 3232, "td_tvd": 3044.07, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-15 D", "well": "F-15", "x": 435053.552, "y": 6478560.355, "td_md": 4685, "td_tvd": 3212.31, "kb": "54.90m", "role": "producer", "has": {"logs": True, "traj": False, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-15", "well": "F-15", "x": 435053.552, "y": 6478560.355, "td_md": 4090, "td_tvd": 3169.92, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-1 A", "well": "F-1", "x": 435046.488, "y": 6478566.687, "td_md": 3682, "td_tvd": 3239.71, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-1 B", "well": "F-1", "x": 435046.488, "y": 6478566.687, "td_md": 3465, "td_tvd": 3259.89, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": False, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-1 C", "well": "F-1", "x": 435046.488, "y": 6478566.687, "td_md": 4094, "td_tvd": 3180.47, "kb": "54.90m", "role": "producer", "has": {"logs": True, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-1", "well": "F-1", "x": 435046.488, "y": 6478566.687, "td_md": 3632, "td_tvd": 3330.4, "kb": "54.90m", "role": "none", "has": {"logs": True, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-4", "well": "F-4", "x": 435049.831, "y": 6478560.825, "td_md": 3510, "td_tvd": 3138.06, "kb": "54.90m", "role": "injector", "has": {"logs": True, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-5", "well": "F-5", "x": 435050.945, "y": 6478558.87, "td_md": 3792, "td_tvd": 3246.43, "kb": "54.90m", "role": "both", "has": {"logs": True, "traj": True, "production": True, "picks": True}, "is_exploration": False},
    {"name": "F-7", "well": "F-7", "x": 435048.907, "y": 6478565.478, "td_md": 1083, "td_tvd": 1077.48, "kb": "54.90m", "role": "none", "has": {"logs": False, "traj": True, "production": False, "picks": True}, "is_exploration": False},
    {"name": "F-9 A", "well": "F-9", "x": 435051.135, "y": 6478561.568, "td_md": 1206, "td_tvd": 1012.97, "kb": "54.90m", "role": "none", "has": {"logs": False, "traj": True, "production": False, "picks": False}, "is_exploration": False},
    {"name": "F-9", "well": "F-9", "x": 435051.135, "y": 6478561.568, "td_md": 1083, "td_tvd": 1075.86, "kb": "54.90m", "role": "none", "has": {"logs": False, "traj": True, "production": False, "picks": True}, "is_exploration": False},
]

# Well = unique surface slot (first occurrence's x/y), Wellbore = every named row
seen_wells = {}
for w in wb_wells:
    seen_wells.setdefault(w["well"], w)

ws = add_sheet("Well")
well_rows = [[well_id(name), FIELD_VOLVE, w["x"], w["y"], "ED50 / UTM 31N", "measured", "C-EQUINOR-VOLVE"] for name, w in seen_wells.items()]
LAST_ROW["Well"] = write_table(ws, ["well_id", "field_id", "x", "y", "crs", "provenance", "source_citation_id"], well_rows,
            nature_col=5, col_widths=[26, 22, 14, 14, 16, 14, 18]) - 1

ws = add_sheet("Wellbore")
wellbore_rows = [
    [wellbore_id(w["name"]), well_id(w["well"]), w["role"], "Y" if w["is_exploration"] else "N", w["td_md"], w["td_tvd"], w["kb"],
     "Y" if w["has"]["logs"] else "N", "Y" if w["has"]["traj"] else "N", "Y" if w["has"]["production"] else "N",
     "Y" if w["has"]["picks"] else "N", "measured", "C-EQUINOR-VOLVE"]
    for w in wb_wells
]
LAST_ROW["Wellbore"] = write_table(ws, ["wellbore_id", "well_id", "role", "is_exploration", "td_md_m", "td_tvd_m", "kb",
                  "has_logs", "has_traj", "has_production", "has_picks", "provenance", "source_citation_id"],
            wellbore_rows, nature_col=11, col_widths=[24, 24, 10, 10, 8, 8, 8, 8, 8, 10, 8, 14, 18]) - 1
note_row(ws, LAST_ROW["Wellbore"] + 2, "Ids use the wellbore NAME as nativeId (authority 'sodir') since only ONE of these 24 has a confirmed real Sodir numeric wellbore id in src/atlas/volve.ts today (well '15/9-F-12', id 5599) — not confidently matched to a specific row here, so not force-fit; if/when confirmed, that row's id can be swapped to atlas:wellbore:sodir:5599 with no schema change.", 13)

# ══════════════════════════════════════════════════════════════════════════════
# COMMERCIAL: Company, Licence, Asset
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Company")
write_table(ws, ["company_id", "name", "role", "country_id", "provenance", "source_citation_id"],
            [[COMPANY_EQUINOR, "Equinor Energy AS", "operator", COUNTRY_NORWAY, "reference", "C-SODIR-VOLVE"]],
            nature_col=4, col_widths=[26, 20, 12, 16, 14, 18])

ws = add_sheet("Licence")
write_table(ws, ["licence_id", "field_id", "operator_id", "status", "provenance", "source_citation_id"],
            [[LICENCE_046BS, FIELD_VOLVE, COMPANY_EQUINOR, "active", "reference", "C-SODIR-VOLVE"]],
            nature_col=4, col_widths=[20, 22, 26, 10, 14, 18])

ws = add_sheet("Asset")
write_table(ws, ["asset_id", "field_id", "fiscal_regime", "provenance", "source_citation_id"],
            [[ASSET_VOLVE, FIELD_VOLVE, "Norway tax + area fee", "reference", "C-SODIR-VOLVE"]],
            nature_col=3, col_widths=[20, 22, 26, 14, 18])

# ══════════════════════════════════════════════════════════════════════════════
# BRIDGE: Study Stages (verbatim from shipped registry.ts), Report Sections
# (generic PBE methodology — NO confidential specifics), Fieldcraft
# ══════════════════════════════════════════════════════════════════════════════
ws = add_sheet("Study Stages")
# verbatim from apps/energy/src/tabs/exploration/registry.ts STUDY_STAGES
study_stages = [
    ["atlas", "Atlas", "Frame", "IHS/S&P - WoodMac - Rystad - USGS", "BasinStats - CreamingCurve - YTFBaseline",
     "Frame the opportunity from world basin evidence and define the study scope.", "untouched"],
    ["data-room", "Data Room", "Frame", "Petrel - GIS", "DataInventory",
     "Inventory seismic, wells and GIS coverage with vintage and quality visible.", "untouched"],
    ["basin-framework", "Basin Framework", "Model", "Neftex - Petrel", "StratColumn - MegaSequence[] - WheelerDiagram",
     "Build the tectonostratigraphic framework that every downstream interpretation references.", "untouched"],
    ["seismic-structure", "Seismic & Structure", "Model", "Petrel - PaleoScan - 2D/3D MOVE",
     "Horizon[] - Fault[] - VelocityModel - Closure[]", "Interpret horizons and faults, depth-convert them and identify closures.", "untouched"],
    ["petrophysics", "Petrophysics", "Model", "Techlog", "ReservoirParameters",
     "Derive evidence-backed PHIE, SWE and NTG distributions by interval.", "untouched"],
    ["gde", "GDE", "Model", "Neftex - SAFARI", "GDEMap[]",
     "Map depositional environments by mega-sequence and expose the basis for each interpretation.", "untouched"],
    ["basin-modeling", "Basin Modeling", "System", "ZetaWare Trinity/T3 - PetroMod - KINEX",
     "MaturityMap - MigrationMap - ChargeTiming - PSEChart",
     "Model burial, maturity, charge timing and screening-grade migration before risking a play.", "untouched"],
    ["play-fairway", "Play Fairway & CRS", "System", "GeoX - Play Chaser - PBE", "PlayDefinition - PDARecord - CRSMap - CCRSMap",
     "Combine charge, reservoir and seal evidence into calibrated play-common risk.", "untouched"],
    ["prospect-risk", "Prospect & Risk", "Decide", "GeoX - Merak Peep", "Opportunity[] - VolumetricCase - GCFAssessment - Ranking",
     "Turn mapped closures into ranked, probabilistic drill-or-drop decisions.", "untouched"],
    ["deliverables", "Deliverables", "Output", "study report - presentation", "StudyReport - StudyPresentation",
     "Compose the approved artifact graph into an auditable report and presentation.", "untouched"],
]
r = write_table(ws, ["stage_id", "name", "phase", "clones", "produces", "blurb", "status"], study_stages,
                col_widths=[16, 20, 10, 30, 40, 55, 12])
LAST_ROW["Study Stages"] = r - 1
note_row(ws, r + 1, "Mirrors apps/energy/src/tabs/exploration/registry.ts STUDY_STAGES, with ONE deliberate change: the 'deliverables' row's clones column reads 'study report - presentation' here instead of the shipped code's literal external-project name, per this workbook's no-confidential-names rule. Keep the rest in sync by hand.", 7)

ws = add_sheet("Report Sections")
report_sections = [
    ["1.0", "Introduction & Study Objectives", "atlas", "Basin; Province", "-", "not built"],
    ["2.0", "Exploration / Regional History", "atlas", "Basin; Field", "Creaming curve", "not built"],
    ["3.0", "Database (seismic & well inventory, QC)", "data-room", "Well; Wellbore", "Data coverage map", "not built"],
    ["4.1", "Regional Geology & Tectonic Framework", "basin-framework", "Basin; Basin Cycle", "Regional cross-section", "not built"],
    ["4.2", "Stratigraphic Framework (mega-sequences)", "basin-framework", "Basin Cycle; Stratigraphic Units", "Well Penetration Chart; Wheeler Diagram; stratigraphic column", "partial — unit data exists, charts not built"],
    ["4.3", "Structural Framework & Seismic Mapping", "seismic-structure", "Play", "Time/Depth Structure Maps; isochrone/isopach", "not built"],
    ["5.0", "Petrophysical Analysis", "petrophysics", "Reservoir; Well", "PHIE/SWE/NTG histograms", "not built"],
    ["6.0", "Gross Depositional Environment (GDE)", "gde", "Basin Cycle; Stratigraphic Units", "GDE map per cycle", "not built"],
    ["7.1", "Source Rock & Hydrocarbon Charge", "basin-modeling", "Petroleum System", "Maturity map; migration ('hairy') map", "not built"],
    ["7.2", "Petroleum System Element Chart / Timing", "basin-modeling", "Petroleum System; PS x Cycle", "PSE chart", "not built"],
    ["7.3", "Play Performance Statistics & Yet-to-Find (YTF)", "basin-modeling", "Basin; Play", "Creaming curve; YTF chart", "not built"],
    ["8.1", "Play Type Identification & Analog Compilation", "play-fairway", "Play; Basin Cycle", "Play type map", "partial — Play row exists, analog compilation not built"],
    ["8.2", "Post-Drilling Analysis (PDA)", "play-fairway", "Well; Field", "PDA 'wagon-wheel' summary", "not built"],
    ["8.3", "CRS (charge / reservoir / seal) & CCRS", "play-fairway", "Play", "CRS maps x3; CCRS map", "not built"],
    ["9.1", "Lead & Prospect Inventory", "prospect-risk", "Opportunity", "Lead/prospect location map", "partial — 1 illustrative Opportunity row exists"],
    ["9.2", "Reservoir Parameter Ranges (P10/P50/P90)", "prospect-risk", "Reservoir; Stratigraphic Units", "Property histograms", "not built"],
    ["9.3", "GRV & Probabilistic Volumetrics", "prospect-risk", "Opportunity", "GRV map; Monte-Carlo histogram", "built — Exploration > Volumetrics tab (apps/energy), verified against wb STOIIP validation"],
    ["9.4", "GCF / Risk Analysis & Resource-vs-GCF Ranking", "prospect-risk", "Opportunity", "Resource-vs-GCF bubble chart; ranking table", "partial — GCF engine (explore.ts) exists, bubble chart not built"],
    ["10.0", "Summary & Recommendation", "deliverables", "ALL", "-", "not built"],
    ["11.0", "References", "deliverables", "Citations", "-", "built — this workbook, Citations tab"],
]
r = write_table(ws, ["section", "title", "stage_id", "required_entity_tabs", "required_figures", "status"], report_sections,
                col_widths=[8, 44, 16, 30, 46, 46])
LAST_ROW["Report Sections"] = r - 1
note_row(ws, r + 1, "GENERIC Play-Based-Exploration report structure (industry-standard PBE workflow terminology). No project name, place name, well name, lead name or confidential number appears here or anywhere in this workbook.", 6)

ws = add_sheet("Fieldcraft Courses")
fc_courses = [
    ["volve-mission", "volve-mission", "The Volve Mission", "From Discovery to Decision", "flagship",
     "Foundation -> Guided Practitioner", "5 days - 40 hours", "Instructor-led; Offline-ready; Enterprise",
     "Integrated lifecycle", 5, 8, "Integrated Geoscience Fieldcraft Passport",
     "Volve; Real data; Geoscience; Field development; Well delivery; Reservoir management"],
    ["exploration-basin-prospect", "exploration-basin-to-prospect", "Exploration Fieldcraft", "Basin to Prospect", "coming-soon",
     "Foundation -> Practitioner", "In development", "Online; Instructor-led",
     "Exploration", 6, 6, "Exploration Vertical Passport",
     "Basin; Petroleum system; Play fairway; Prospect risk; Volumetrics"],
]
LAST_ROW["Fieldcraft Courses"] = write_table(ws, ["course_id", "slug", "title", "subtitle", "status", "level", "duration", "delivery", "lifecycle",
                  "modules", "labs", "credential", "tags"], fc_courses,
            col_widths=[24, 24, 20, 24, 12, 22, 16, 30, 16, 8, 6, 32, 46]) - 1

ws = add_sheet("Fieldcraft Course Days")
fc_days = [
    [1, "volve-mission", "DISCOVER", "Frame the opportunity", "Exploration", "exploration",
     "What do we know, how do we know it, and is the opportunity mature enough to progress?",
     "Exploration Gate Card - Progress, Study or Stop", BASIN_VIKING, FIELD_VOLVE,"19 A; 19 BT2; 19 SR (exploration wells)"],
    [2, "volve-mission", "DESCRIBE & DESIGN", "Build the field case", "Field Development", "field-development",
     "Which subsurface and development case is technically defensible and decision-relevant?",
     "Development Case Card - Select, Rework or Reject", BASIN_VIKING, FIELD_VOLVE,"F-11; F-12; F-14; F-15 D; F-1 C (producers)"],
    [3, "volve-mission", "DELIVER", "Turn the plan into a well", "Well Delivery & Drilling", "well-delivery",
     "Is the proposed well technically ready, safely framed and schedulable?",
     "Well Gate Card - Approve, Condition or Hold", BASIN_VIKING, FIELD_VOLVE,"F-12 (representative candidate well)"],
    [4, "volve-mission", "OPERATE", "Monitor, diagnose and act", "Reservoir Management", "reservoir-management",
     "Which field signal matters, what is the defensible diagnosis and what should happen next?",
     "Reservoir Action Card - Act, Acquire Data or Monitor", BASIN_VIKING, FIELD_VOLVE,"F-4; F-5; F-11; F-12; F-14; F-15 D (producers + injectors)"],
    [5, "volve-mission", "DECIDE", "Integrated Field Review", "Cross-lifecycle", "cockpit",
     "Considering the complete evidence trail, what is the next best field decision and why?",
     "Integrated Field Decision - Fieldcraft Passport", BASIN_VIKING, FIELD_VOLVE,"all wells (cross-lifecycle capstone)"],
]
r = write_table(ws, ["day_number", "course_id", "verb", "title", "lifecycle", "workspace", "question", "outcome",
                      "uses_basin_id", "uses_field_id", "uses_wellbore_ids"], fc_days,
                col_widths=[10, 16, 18, 24, 20, 16, 46, 40, 14, 12, 46])
LAST_ROW["Fieldcraft Course Days"] = r - 1
note_row(ws, r + 1, "day_number..outcome mirror apps/energy/src/fieldcraft/catalog.ts VOLVE_DAYS verbatim. uses_basin_id/uses_field_id/uses_wellbore_ids are a NEW proposed bridge (not yet in the shipped CourseDay type) — SUGGESTED linkage inferred from each day's workspace, not yet wired into the app. Full schedule/slides/materials content lives in catalog.ts, not duplicated here to avoid drift.", 11)

# ══════════════════════════════════════════════════════════════════════════════
# README LIVE COUNTS — written last, now that every sheet's exact last-data-row is
# known. Bounded COUNTA(Sheet!A2:A{last_row}) ranges — NOT whole-column A:A — so a
# note row placed below any table is never miscounted as data.
# ══════════════════════════════════════════════════════════════════════════════
readme = wb["README"]
count_labels = [
    ("Regions", "Region"), ("Countries", "Country"),
    ("Provinces", "Province"), ("Basins", "Basin"), ("Basin cycles", "Basin Cycle"),
    ("Stratigraphic units", "Stratigraphic Units"), ("Petroleum systems", "Petroleum System"),
    ("Assessment units", "Assessment Unit"), ("Plays", "Play"), ("Opportunities (leads/prospects)", "Opportunity"),
    ("Fields", "Field"), ("Reservoirs", "Reservoir"), ("Wells", "Well"), ("Wellbores", "Wellbore"),
    ("Citations", "Citations"), ("Concepts", "Concepts"), ("Doust figures sourced", "Doust Figure Sourcing"),
    ("Study stages", "Study Stages"), ("Report sections", "Report Sections"),
    ("Fieldcraft courses", "Fieldcraft Courses"), ("Fieldcraft course days", "Fieldcraft Course Days"),
]
rr = LIVE_COUNTS_ROW
missing = [s for _, s in count_labels if s not in LAST_ROW]
assert not missing, f"LAST_ROW missing entries for: {missing}"
for label, sheet in count_labels:
    readme.cell(row=rr, column=1, value=label).font = BODY_FONT
    f = readme.cell(row=rr, column=2, value=f"=COUNTA('{sheet}'!A2:A{LAST_ROW[sheet]})")
    f.font = BODY_FONT
    rr += 1
note_row(readme, rr + 1, "Bounded ranges (A2:A{last real data row}) — deliberately NOT whole-column A:A, so explanatory note rows below a table are never counted as data. A tab intentionally left empty (Block-Licence, Pool) is not on this list.", 2)

wb.save(OUT)
print("workbook complete —", len(wb.sheetnames), "sheets")
print("LAST_ROW captured for:", sorted(LAST_ROW.keys()))
