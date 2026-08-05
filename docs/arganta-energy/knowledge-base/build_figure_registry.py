# build_figure_registry.py — Phase 0 of the figure-governance programme.
#
# Turns the flat basin-figure manifest into two governed sheets:
#
#   Figure Registry — ONE row per figure. The figure is the evidence object; it exists
#                     independently of who uses it.
#   Figure Links    — junction. One figure may serve many formations, basins, TPS or
#                     fields, each with its own relationship and relevance rank.
#
# Why the junction matters: the manifest carries 703 entries over only 325 images,
# because a figure shared across provinces was duplicated once per province. That is
# the exact shape a junction table exists to fix — one regional depositional model can
# serve eight formations without eight near-copies of the same row.
#
# RIGHTS MODEL — the important change here.
# The old classifier returned 'usgs-public-domain' whenever it found no credit line.
# That fails OPEN: anything unrecognised was assumed safe. Defensible while the corpus
# was USGS-only; wrong the moment anything else is ingested. Licence and redistribution
# are now SEPARATE fields, and the default is DO-NOT-INGEST.
#
# Run: python docs/arganta-energy/knowledge-base/build_figure_registry.py
import json
import os
import re
import shutil
import time
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment

ROOT = r"C:\Users\aldhy\OneDrive\Documents\GitHub\ArgantaLab"
XLSX = os.path.join(ROOT, "docs", "arganta-energy", "knowledge-base", "ArgantaEnergy-Master-KB.xlsx")
MANIFEST = os.path.join(ROOT, "apps", "energy", "public", "basin-figures", "manifest.json")
SPINE = os.path.join(ROOT, "apps", "energy", "public", "kb", "master-kb-spine.json")
TMP = os.path.join(ROOT, ".codex", "tmp-petsys", "_figreg.xlsx")

REGISTRY_COLS = [
    "figure_id", "title", "figure_scope", "figure_type",
    "formation_id", "basin_id", "tps_id", "field_id",
    "source_citation_id", "source_url", "doi", "publication_year", "page", "figure_number",
    "caption", "authority_type", "geographic_scope", "age_scope", "content_summary",
    "decision_use", "resolution_quality", "scientific_quality",
    "licence_status", "redistribution_status", "local_asset_path", "thumbnail_allowed",
    "candidate_score", "score_coverage_pct", "preferred_for_scope",
    "superseded_by", "review_status", "reviewer_notes",
]
# preferred_for_scope lives on the LINK, not the figure. "Preferred" is a judgement
# about a figure's fitness FOR A GIVEN ENTITY: the same regional chart can be the
# preferred general chart for one basin and merely an alternate for its neighbour.
# Holding it on the figure forces one global winner and silently loses the rest.
LINK_COLS = ["figure_link_id", "figure_id", "entity_type", "entity_id",
             "relationship", "relevance_rank", "preferred_for_scope", "notes"]

# ── vocabularies (closed; the validator enforces them) ───────────────────────
LICENCE = ["public-domain", "cc-by", "cc-by-sa", "cc-by-nc", "all-rights-reserved", "unknown"]
REDIST = ["local-copy-permitted", "link-only", "internal-reference-only", "do-not-ingest"]
AUTHORITY = ["geological-survey", "petroleum-regulator", "peer-reviewed", "basin-atlas",
             "operator", "textbook", "unknown"]

# Petroleum-system relevance by figure type — drives 15% of the candidate score.
PS_RELEVANCE = {
    "ps-summary": 1.0, "events-chart": 1.0, "kitchen-migration-map": 0.95,
    "burial-history": 0.9, "oil-source-correlation": 0.9, "source-rock-panel": 0.9,
    "seal-capacity": 0.85, "cross-section": 0.8, "play-fairway": 0.8,
    "strat-chart": 0.7, "depositional": 0.65, "facies-architecture": 0.6,
    "poro-perm": 0.6, "basin-evolution": 0.6, "paleogeographic": 0.55,
    "correlation-panel": 0.5, "type-log": 0.5, "map": 0.4, "creaming": 0.35,
    "other": 0.2,
}
# Formation-level resolution implied by the type — 15% of the score.
FORMATION_RES = {
    "strat-chart": 1.0, "correlation-panel": 1.0, "type-log": 1.0,
    "facies-architecture": 0.9, "depositional": 0.7, "cross-section": 0.6,
    "source-rock-panel": 0.8, "poro-perm": 0.8, "seal-capacity": 0.8,
    "events-chart": 0.5, "paleogeographic": 0.4, "map": 0.15, "creaming": 0.1,
    "other": 0.1,
}
WEIGHTS = {  # from the handoff
    "authority": 0.20, "basin_match": 0.15, "formation_res": 0.15, "age_clarity": 0.10,
    "ps_relevance": 0.15, "coverage": 0.10, "recency": 0.05, "legibility": 0.05,
    "reuse": 0.05,
}


def classify_rights(rights, credit):
    """Map the harvester's coarse flag onto the two governed fields.

    Licence describes what the rightsholder allows; redistribution describes what WE
    may do with it. Conflating them is how a 'we have a local copy' fact silently
    becomes a 'we may publish it' claim.
    """
    if rights == "usgs-public-domain":
        return "public-domain", "local-copy-permitted", "yes", "geological-survey"
    if rights == "cc-attribution":
        return "cc-by", "local-copy-permitted", "yes", "peer-reviewed"
    if rights == "restricted":
        # Third-party figure reproduced inside a public-domain report. We hold a local
        # copy for reference; we may not redistribute it, and not even as a thumbnail
        # unless the rightsholder says so.
        return "all-rights-reserved", "internal-reference-only", "unknown", "peer-reviewed"
    # FAIL CLOSED. Anything unrecognised is quarantined, not assumed safe.
    return "unknown", "do-not-ingest", "no", "unknown"


YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def year_of(pub):
    m = YEAR_RE.search(pub or "")
    return int(m.group(0)) if m else None


def score(fig, shared):
    """Weighted candidate score over the criteria we can actually assess.

    Several inputs in the handoff's rubric — scientific quality, coverage, age clarity —
    need a human or a full read of the source. Scoring them as 0 would silently punish
    every figure; scoring them as 1 would inflate every figure. Instead the score is a
    weighted average over the ASSESSED criteria only, and score_coverage_pct records how
    much of the rubric that represents, so nobody reads 0.82 as a complete judgement.
    """
    got, total = 0.0, 0.0
    ftype = fig["type"]

    def add(key, value):
        nonlocal got, total
        got += WEIGHTS[key] * value
        total += WEIGHTS[key]

    add("authority", 0.9 if fig["rights"] != "unknown" else 0.3)   # USGS = surveyed + reviewed
    add("basin_match", 0.6 if shared else 1.0)                     # shared multi-province = looser
    add("formation_res", FORMATION_RES.get(ftype, 0.1))
    add("ps_relevance", PS_RELEVANCE.get(ftype, 0.2))
    yr = year_of(fig.get("source_publication"))
    if yr:
        add("recency", max(0.0, min(1.0, (yr - 1990) / 30)))
    px = (fig.get("w") or 0) * (fig.get("h") or 0)
    add("legibility", max(0.0, min(1.0, px / 600_000)))            # ~1000x600 = full marks
    add("reuse", {"usgs-public-domain": 1.0, "cc-attribution": 0.8,
                  "restricted": 0.2}.get(fig["rights"], 0.0))
    # age_clarity, coverage, scientific_quality deliberately NOT scored — they need the
    # source read, which is Phase 4 work.
    return round(got / total, 3) if total else 0.0, round(total * 100)


def main():
    man = json.load(open(MANIFEST, encoding="utf-8"))
    kb = json.load(open(SPINE, encoding="utf-8"))
    basin_name = {b["basin_id"]: b.get("name") for b in kb["basin"]}
    figs = man["figures"]

    # collapse manifest entries onto their unique image — that IS the figure
    by_file = defaultdict(list)
    for f in figs:
        by_file[f["file"]].append(f)
    print(f"manifest entries {len(figs)} → unique figures {len(by_file)}")

    registry, links = [], []
    for file, entries in sorted(by_file.items()):
        head = entries[0]
        shared = len(entries) > 1
        lic, redist, thumb, authority = classify_rights(head["rights"], head.get("credit"))
        sc, cov = score(head, shared)
        fid = "atlas:figure:" + re.sub(r"\.png$", "", file)
        pub = head.get("source_publication") or ""
        registry.append({
            "figure_id": fid,
            "title": (head.get("caption") or "")[:180],
            # every figure harvested so far is basin/province scoped; formation scoping
            # arrives in Phase 1 when formations become real entities
            "figure_scope": "basin",
            "figure_type": head["type"],
            "formation_id": None,
            "basin_id": head.get("basin_id"),
            "tps_id": None,
            "field_id": None,
            "source_citation_id": "C-USGS-DDS69" if "DDS-60" not in pub else "C-USGS-DDS60-AU",
            "source_url": None,
            "doi": None,
            "publication_year": year_of(pub),
            "page": head.get("source_page"),
            "figure_number": (re.search(r"fig(\d+)", file).group(1) if re.search(r"fig(\d+)", file) else None),
            "caption": head.get("caption"),
            "authority_type": authority,
            "geographic_scope": basin_name.get(head.get("basin_id")),
            "age_scope": None,
            "content_summary": None,
            "decision_use": None,
            "resolution_quality": f"{head.get('w')}x{head.get('h')}",
            "scientific_quality": None,
            "licence_status": lic,
            "redistribution_status": redist,
            "local_asset_path": (("apps/energy/public/basin-figures-restricted/" if head["restricted"]
                                  else "apps/energy/public/basin-figures/") + file),
            "thumbnail_allowed": thumb,
            "candidate_score": sc,
            "score_coverage_pct": cov,
            "preferred_for_scope": None,
            "superseded_by": None,
            "review_status": "auto-classified",
            "reviewer_notes": ("Scored on the assessable criteria only; age clarity, coverage and "
                               "scientific quality require the source to be read (Phase 4)."
                               + (" Shared across multiple provinces." if shared else "")),
        })
        for i, e in enumerate(sorted(entries, key=lambda x: x["id"])):
            if not e.get("basin_id"):
                continue
            links.append({
                "figure_link_id": f"atlas:figure-link:{re.sub(r'.png$', '', file)}:{i+1}",
                "figure_id": fid,
                "entity_type": "basin",
                "entity_id": e["basin_id"],
                "relationship": "depicts",
                "relevance_rank": 1 if not shared else i + 1,
                "preferred_for_scope": None,
                "notes": None if not shared else "figure published for a multi-province assessment",
            })

    # ── preferred_for_scope — per (entity, purpose) ─────────────────────────
    # A basin's "preferred general" chart and its "preferred petroleum-system" chart are
    # different questions with different winners, so each slot is resolved separately
    # and a figure may legitimately win several. Only figures we may actually show are
    # eligible: an internal-reference-only plate must never become a basin's default.
    reg_by_id = {r["figure_id"]: r for r in registry}
    links_by_entity = defaultdict(list)
    for l in links:
        if reg_by_id[l["figure_id"]]["redistribution_status"] == "local-copy-permitted":
            links_by_entity[l["entity_id"]].append(l)
    SLOTS = {
        "preferred_general": lambda r: r["candidate_score"],
        "preferred_petroleum_system": lambda r: PS_RELEVANCE.get(r["figure_type"], 0) * r["candidate_score"],
        "preferred_high_resolution": lambda r: (int(r["resolution_quality"].split("x")[0])
                                                * int(r["resolution_quality"].split("x")[1])),
    }
    for entity, ls in links_by_entity.items():
        for slot, key in SLOTS.items():
            best = max(ls, key=lambda l: key(reg_by_id[l["figure_id"]]), default=None)
            if best:
                best["preferred_for_scope"] = ((best["preferred_for_scope"] + ";" + slot)
                                               if best["preferred_for_scope"] else slot)
    for l in links:
        if not l["preferred_for_scope"]:
            l["preferred_for_scope"] = "alternate"
    # On the figure itself the field records its own best ROLE, not a per-basin verdict.
    for r in registry:
        r["preferred_for_scope"] = None

    # ── write ───────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(XLSX)
    for name, cols, rows in (("Figure Registry", REGISTRY_COLS, registry),
                             ("Figure Links", LINK_COLS, links)):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        hf, ha = Font(bold=True), Alignment(vertical="top", wrap_text=True)
        for i, c in enumerate(cols):
            cell = ws.cell(row=1, column=i + 1)
            cell.value = c
            cell.font, cell.alignment = hf, ha
        for r_i, row in enumerate(rows):
            for c_i, c in enumerate(cols):
                ws.cell(row=r_i + 2, column=c_i + 1).value = row.get(c)
        ws.freeze_panes = "A2"
        print(f"  {name}: {len(rows)} rows x {len(cols)} cols")

    wb.save(TMP)
    for i in range(6):
        try:
            shutil.move(TMP, XLSX); print("SAVED."); break
        except OSError as e:
            print("  retry", i, e.errno); time.sleep(2)
    else:
        raise SystemExit("could not move into place")

    from collections import Counter
    print()
    print("licence:", dict(Counter(r["licence_status"] for r in registry)))
    print("redistribution:", dict(Counter(r["redistribution_status"] for r in registry)))
    slots = Counter(x for l in links for x in (l["preferred_for_scope"] or "").split(";") if x)
    print("preferred slots (on links):", dict(slots))
    print("basins with a preferred_general:",
          len({l["entity_id"] for l in links if "preferred_general" in (l["preferred_for_scope"] or "")}))
    print(f"score range: {min(r['candidate_score'] for r in registry)}"
          f" – {max(r['candidate_score'] for r in registry)}"
          f"  (rubric coverage {registry[0]['score_coverage_pct']}%)")


if __name__ == "__main__":
    main()
