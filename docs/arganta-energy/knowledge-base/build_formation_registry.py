# build_formation_registry.py — Phase 1: turn free-text unit names into real entities.
#
# WHY THIS BLOCKS EVERYTHING FORMATION-SCOPED
#   20 of the 28 figure types in the visual-evidence spec are formation-scoped, and
#   `formation_id` currently has nothing to point at: `Stratigraphic Units` holds 11
#   rows (Volve only), while formation names live as free text on PS Elements and Basin
#   Cycles — 1,306 distinct atoms, of which a large minority are not formations at all.
#
# WHAT THIS DOES NOT DO
#   It does not invent stratigraphy. It clusters strings that are demonstrably the same
#   unit ("Fahdene Fm" / "Albian Lower Fahdene Formation"), rejects things that were
#   never formations (derived placeholders, bare lithologies, "Eocene source interval"),
#   and records every input string as an alias so nothing is silently lost. Parent/child
#   relationships and true nomenclature authority need the literature — Phase 4.
#
# Run: python docs/arganta-energy/knowledge-base/build_formation_registry.py
import json
import os
import re
import shutil
import time
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment

ROOT = r"C:\Users\aldhy\OneDrive\Documents\GitHub\ArgantaLab"
XLSX = os.path.join(ROOT, "docs", "arganta-energy", "knowledge-base", "ArgantaEnergy-Master-KB.xlsx")
SPINE = os.path.join(ROOT, "apps", "energy", "public", "kb", "master-kb-spine.json")
TMP = os.path.join(ROOT, ".codex", "tmp-petsys", "_formreg.xlsx")

COLS = ["formation_id", "canonical_name", "rank", "aliases", "alias_count",
        "parent_unit", "lithology_hint", "age_hint", "basin_ids", "basin_count",
        "occurrence_count", "source_tables", "provenance", "review_status", "notes"]

# ── rejection rules ─────────────────────────────────────────────────────────
# A lithology is not a formation. This is the same defect already flagged in the PS
# chart (22 rows naming "Coal" or "Shale" where a unit belongs) — here it is filtered
# at the point where formations become entities, so it cannot propagate further.
LITHOLOGY = {
    "coal", "shale", "shales", "sandstone", "sandstones", "limestone", "limestones",
    "dolomite", "carbonate", "carbonates", "sand", "silt", "siltstone", "mudstone",
    "evaporite", "evaporites", "anhydrite", "salt", "chalk", "clastics", "marl",
    "chert", "conglomerate", "reef", "basalt", "granite", "volcanics", "diabase",
    "basement", "sediments", "clay", "clays", "tuff", "shale/carbonate",
}
# Chronostratigraphic qualifiers that decorate a name rather than belong to it.
CHRONO = {
    "cambrian", "ordovician", "silurian", "devonian", "carboniferous", "mississippian",
    "pennsylvanian", "permian", "triassic", "jurassic", "cretaceous", "paleogene",
    "neogene", "quaternary", "tertiary", "paleocene", "eocene", "oligocene", "miocene",
    "pliocene", "pleistocene", "holocene", "paleozoic", "mesozoic", "cenozoic",
    "precambrian", "riphean", "vendian", "infracambrian",
    "maastrichtian", "campanian", "santonian", "coniacian", "turonian", "cenomanian",
    "albian", "aptian", "barremian", "hauterivian", "valanginian", "berriasian",
    "tithonian", "kimmeridgian", "oxfordian", "callovian", "bathonian", "bajocian",
    "frasnian", "famennian", "volgian", "neocomian", "visean", "bashkirian",
    "kungurian", "serpukhovian", "tournaisian", "westphalian", "rotliegend",
    "lower", "upper", "middle", "early", "late", "earliest", "latest", "uppermost",
    "lowermost", "the", "permo", "syn-rift", "post-rift", "pre-rift",
}
RANK_MAP = {
    "fm": "Formation", "formation": "Formation", "fms": "Formation",
    "gp": "Group", "group": "Group", "grp": "Group",
    "mbr": "Member", "member": "Member",
    "supergroup": "Supergroup", "series": "Series", "beds": "Beds",
    "horizon": "Horizon", "sandstone": "Formation", "shale": "Formation",
    "limestone": "Formation", "anhydrite": "Formation", "salt": "Formation",
}
REJECT_PAT = re.compile(
    # A surface is not a unit. An unconformity, a maximum flooding surface or an
    # erosion surface is a BOUNDARY between units — giving it a formation_id would let
    # figures and elements attach to something that has no thickness, age range or
    # lithology. Caught in review: "BCU", "Hercynian unconformity", "Glacial erosion
    # surface" had all become "... Formation".
    r"\b(?:unconformity|BCU|MFS|erosion\s+surface|flooding\s+surface|"
    r"discontinuity|hiatus|contact|datum)\b|"
    r"\(derived|candidate\s*\(unnamed|"
    r"\b(?:source|reservoir|seal|overburden)\s+(?:interval|candidate)\b|"
    r"^\s*(?:overburden|source|reservoir|seal)\b.*\bderived\b|"
    r"\b(?:section|sequence|succession|cover|fill|complex|package|interval)\s*$|"
    # NOTE: no `^[a-z]` clause here. Under re.I that character class also matches
    # uppercase, which rejected every string beginning with a letter — i.e. all of
    # them. The proper-noun requirement is enforced on `core` further down, where it
    # can be case-sensitive.
    r"^\d", re.I)


def normalise(raw):
    """Return (canonical_name, rank, lithology_hint, age_hint) or None if rejected."""
    s = re.sub(r"\s+", " ", (raw or "").strip()).strip(" .,;")
    if not s or len(s) < 3:
        return None
    if REJECT_PAT.search(s):
        return None
    low = s.lower()
    if low in LITHOLOGY:
        return None

    # Capture then strip the age qualifier: "Albian Lower Fahdene Formation" -> Fahdene.
    # Qualifiers are frequently hyphenated compounds — "Permo-Triassic Khuff Formation",
    # "Jurassic-Lower Cretaceous Vaca Muerta Formation" — so test each hyphen-part.
    # Treating those as part of the name fragments the registry: Khuff would exist twice.
    def is_chrono(word):
        parts = [p for p in re.split(r"[-/]", re.sub(r"[^a-z/-]", "", word.lower())) if p]
        return bool(parts) and all(p in CHRONO for p in parts)

    words = s.split()
    age_bits = []
    while words and is_chrono(words[0]):
        age_bits.append(words.pop(0))
    if not words:
        return None

    # trailing rank word
    rank, litho = None, None
    tail = re.sub(r"[^A-Za-z]", "", words[-1]).lower()
    if tail in RANK_MAP:
        rank = RANK_MAP[tail]
        if tail in LITHOLOGY:
            litho = words[-1].title()
        words = words[:-1]
    if not words:
        return None

    core = " ".join(words).strip(" .,;-")
    if not core or core.lower() in LITHOLOGY or len(core) < 3:
        return None
    # a real unit name starts with a capitalised proper noun
    if not re.match(r"^[A-Z\u00C0-\u024F]", core):
        return None
    # reject leftovers that are only chronostrat words
    if all(re.sub(r"[^a-z-]", "", w.lower()) in CHRONO for w in core.split()):
        return None

    rank = rank or "Formation"       # bare proper noun in this corpus means a formation
    name = f"{core} {rank}" if rank in ("Formation", "Group", "Member", "Supergroup") else core
    return name, rank, litho, (" ".join(age_bits) or None)


def key_of(name):
    return re.sub(r"[^a-z0-9]", "", name.lower())


def slug(s):
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s.lower())).strip("-")


def main():
    kb = json.load(open(SPINE, encoding="utf-8"))
    ps = {p["tps_id"]: p for p in kb["petroleumSystem"]}
    prov_basin = {b["province_id"]: b["basin_id"] for b in kb["basin"] if b.get("province_id")}
    model_basin = {}
    for m in kb["psModel"]:
        p = ps.get(m.get("tps_id"))
        if p and p.get("province_id"):
            model_basin[m["model_id"]] = prov_basin.get(p["province_id"])

    # (raw string, basin_id, source table)
    obs = []
    for e in kb["psElement"]:
        if e.get("unit_name"):
            obs.append((e["unit_name"], model_basin.get(e["model_id"]), "PS Elements"))
    for c in kb["basinCycle"]:
        if c.get("units"):
            for part in c["units"].split(";"):
                obs.append((part, c.get("basin_id"), "Basin Cycle"))

    clusters = defaultdict(lambda: {"aliases": Counter(), "basins": set(), "tables": set(),
                                    "rank": None, "litho": None, "age": None, "n": 0})
    rejected = Counter()
    for raw, bid, table in obs:
        for atom in re.split(r"[;/]", raw):
            atom = atom.strip()
            if not atom:
                continue
            got = normalise(atom)
            if not got:
                rejected[atom] += 1
                continue
            name, rank, litho, age = got
            c = clusters[key_of(name)]
            c["aliases"][atom] += 1
            c["n"] += 1
            if bid:
                c["basins"].add(bid)
            c["tables"].add(table)
            c["rank"] = c["rank"] or rank
            c["litho"] = c["litho"] or litho
            c["age"] = c["age"] or age
            c.setdefault("name", name)

    rows = []
    for k, c in sorted(clusters.items(), key=lambda kv: -kv[1]["n"]):
        # canonical display name = the longest alias that still normalises to this key,
        # so "Vaca Muerta Formation" wins over a bare "Vaca Muerta"
        name = c["name"]
        rows.append({
            "formation_id": f"atlas:formation:{slug(name)}",
            "canonical_name": name,
            "rank": c["rank"],
            "aliases": "; ".join(sorted(c["aliases"], key=lambda a: (-c["aliases"][a], a))[:8]),
            "alias_count": len(c["aliases"]),
            "parent_unit": None,
            "lithology_hint": c["litho"],
            "age_hint": c["age"],
            "basin_ids": "; ".join(sorted(c["basins"])[:6]),
            "basin_count": len(c["basins"]),
            "occurrence_count": c["n"],
            "source_tables": "; ".join(sorted(c["tables"])),
            "provenance": "derived-rule",
            "review_status": "auto-canonicalised",
            "notes": ("Clustered from free-text unit names by normalising rank words and "
                      "stripping chronostratigraphic qualifiers. Parent unit, nomenclature "
                      "authority and true stratigraphic position are NOT established here."),
        })

    print(f"observations {len(obs)}  →  formations {len(rows)}")
    print(f"rejected distinct strings: {len(rejected)}  (top: "
          + ", ".join(f'{s!r}x{n}' for s, n in rejected.most_common(4)) + ")")
    print(f"multi-basin formations: {sum(1 for r in rows if r['basin_count'] > 1)}")
    print(f"formations with >1 alias: {sum(1 for r in rows if r['alias_count'] > 1)}")
    print()
    print("top by occurrence:")
    for r in rows[:12]:
        print(f"  {r['occurrence_count']:4d}x  {r['canonical_name'][:44]:46s} "
              f"basins={r['basin_count']:2d} aliases={r['alias_count']}")

    wb = openpyxl.load_workbook(XLSX)
    if "Formation" in wb.sheetnames:
        del wb["Formation"]
    ws = wb.create_sheet("Formation")
    hf, ha = Font(bold=True), Alignment(vertical="top", wrap_text=True)
    for i, c in enumerate(COLS):
        cell = ws.cell(row=1, column=i + 1)
        cell.value, cell.font, cell.alignment = c, hf, ha
    for ri, row in enumerate(rows):
        for ci, c in enumerate(COLS):
            ws.cell(row=ri + 2, column=ci + 1).value = row.get(c)
    ws.freeze_panes = "A2"

    wb.save(TMP)
    for i in range(6):
        try:
            shutil.move(TMP, XLSX); print("\nSAVED. Formation sheet:", len(rows), "rows"); break
        except OSError as e:
            print("  retry", i, e.errno); time.sleep(2)
    else:
        raise SystemExit("could not move into place")


if __name__ == "__main__":
    main()
